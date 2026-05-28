import streamlit as st
import sqlite3
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, date
import logging

from config import DB_FILE, PRODUCTION_MODE
from core.pricing_engine import PricingEngine, PricingInput
from core.hierarchy_resolver import HierarchyResolver

logging.basicConfig(level=logging.WARNING)

st.set_page_config(page_title="Bunker Commerciale - Salov", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
    :root { color-scheme: light !important; }
    .stApp { background-color: #F8F9FA !important; } 
    
    section[data-testid="stSidebar"] { 
        background-color: #FFFFFF !important; 
        border-right: 1px solid #E0E0E0 !important; 
        box-shadow: 2px 0 5px rgba(0,0,0,0.02);
    }
    
    h1, h2, h3, h4, h5, h6 { 
        color: #1A3E2F !important; 
        font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif !important;
        font-weight: 700 !important; 
        letter-spacing: -0.5px;
    }
    .stMarkdown p, .stMarkdown li, label { 
        color: #333333 !important; 
        font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif !important; 
        font-size: 1rem !important; 
    }

    div[data-testid="stMetric"] {
        background-color: #FFFFFF;
        padding: 15px 20px;
        border-radius: 8px;
        border: 1px solid #EAEAEA;
        box-shadow: 0 2px 4px rgba(0,0,0,0.03);
    }
    div[data-testid="stMetricValue"] { 
        color: #D32F2F !important; 
        font-weight: 800 !important; 
        font-size: 1.8rem !important;
    }
    div[data-testid="stMetricLabel"] {
        font-size: 0.9rem !important;
        color: #666666 !important;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    div[data-testid="stVerticalBlock"] > div[style*="border"] { 
        background-color: #FFFFFF !important;
        border-radius: 10px !important; 
        border: 1px solid #E0E0E0 !important; 
        box-shadow: 0 4px 6px rgba(0,0,0,0.04) !important; 
        padding: 20px !important;
    }

    div[data-testid="stExpander"] { 
        background-color: #FFFFFF !important; 
        border: 1px solid #D1C9BC !important; 
        border-radius: 8px !important; 
        overflow: hidden;
    }
    div[data-testid="stExpander"] summary {
        background-color: #F5F7F5 !important; 
        font-weight: 600;
        color: #1A3E2F;
    }

    .stButton>button {
        background-color: #1A3E2F !important;
        color: #FFFFFF !important;
        font-weight: 600 !important;
        border-radius: 6px !important;
        border: none !important;
        padding: 10px 24px !important;
        transition: all 0.2s ease-in-out;
    }
    .stButton>button:hover {
        background-color: #2E7D32 !important;
        box-shadow: 0 4px 8px rgba(46, 125, 50, 0.3) !important;
        transform: translateY(-1px);
    }

    div[data-testid="stDataFrame"] {
        border-radius: 8px;
        border: 1px solid #E0E0E0;
        overflow: hidden;
    }
    
    /* --- FIX FORZATO PER LA FRECCIA DELLA SIDEBAR --- */
    header[data-testid="stHeader"] {
        visibility: visible !important;
        display: block !important;
        opacity: 1 !important;
        z-index: 999998 !important;
        background: transparent !important;
    }
    
    [data-testid="collapsedControl"] {
        display: flex !important;
        visibility: visible !important;
        opacity: 1 !important;
        z-index: 999999 !important;
        color: #1A3E2F !important; 
        background-color: #E8F5E9 !important; 
        border-radius: 6px !important;
        box-shadow: 0 2px 4px rgba(0,0,0,0.2) !important;
        margin-top: 10px !important;
        margin-left: 10px !important;
    }
    
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

def init_db():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='anagrafica_master'")
    db_inizializzato = cursor.fetchone()
    
    if not db_inizializzato:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS anagrafica_master (
            ean TEXT PRIMARY KEY, codice_sap TEXT, tipo_olio TEXT,
            descrizione_sap TEXT, descrizione_commerciale TEXT, formato_lt REAL,
            confezione TEXT, pezzi_cartone INTEGER, cartoni_strato INTEGER,
            strati_pallet INTEGER, cartoni_pallet INTEGER, conservazione_mesi INTEGER, shelf_life_mesi INTEGER
        )""")
        
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS guardrail_aziendali (
            ean TEXT PRIMARY KEY, min_net_net_g REAL DEFAULT 0.0
        )""")
        
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS clienti (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            gruppo_macro TEXT, sottogruppo TEXT, associato_insegna TEXT,
            attivo BOOLEAN DEFAULT 1, UNIQUE(gruppo_macro, sottogruppo, associato_insegna)
        )""")
        
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS accordi_commerciali (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            gruppo_macro TEXT, sottogruppo TEXT, associato_insegna TEXT, livello TEXT, chiave_livello TEXT,
            listino_r REAL, sconto_1 REAL, sconto_2 REAL, sconto_3 REAL, sconto_4 REAL, sconto_5 REAL,
            sconto_6 REAL, sconto_7 REAL, sconto_y REAL, sconto_carico REAL, sconto_pagamento REAL,
            voce_contratto_1 REAL, voce_contratto_2 REAL, voce_contratto_3 REAL, voce_contratto_4 REAL, voce_contratto_5 REAL,
            UNIQUE(gruppo_macro, sottogruppo, associato_insegna, livello, chiave_livello)
        )""")
        
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS storico_promo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data_salvataggio TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            stato_promo TEXT,
            gruppo_macro TEXT, sottogruppo TEXT, associato_insegna TEXT,
            ean TEXT, descrizione_commerciale TEXT,
            listino_r REAL, sconto_y REAL, sconto_z REAL, sconto_aa REAL,
            net_net_am REAL,
            volumi_stimati INTEGER, contributo_fisso REAL, contributo_pezzo REAL, costo_totale_extra REAL,
            note TEXT,
            sell_in_dal DATE, sell_in_al DATE, sell_out_dal DATE, sell_out_al DATE,
            min_net_net_g REAL, net_net_post_promo REAL
        )""")
        conn.commit()
        seed_baseline_data(conn)
    else:
        # Migrazione automatica per aggiungere le nuove colonne se la tabella esiste già
        try:
            cursor.execute("ALTER TABLE storico_promo ADD COLUMN sell_in_dal DATE")
            cursor.execute("ALTER TABLE storico_promo ADD COLUMN sell_in_al DATE")
            cursor.execute("ALTER TABLE storico_promo ADD COLUMN sell_out_dal DATE")
            cursor.execute("ALTER TABLE storico_promo ADD COLUMN sell_out_al DATE")
            cursor.execute("ALTER TABLE storico_promo ADD COLUMN min_net_net_g REAL")
            cursor.execute("ALTER TABLE storico_promo ADD COLUMN net_net_post_promo REAL")
            conn.commit()
        except sqlite3.OperationalError:
            pass # Le colonne esistono già
            
    conn.close()
    
def seed_baseline_data(conn):
    cursor = conn.cursor()
    cursor.execute("DELETE FROM accordi_commerciali")
    cursor.execute("DELETE FROM clienti")
    cursor.execute("DELETE FROM anagrafica_master")
    cursor.execute("DELETE FROM guardrail_aziendali")
    
    prodotti_salov = [
        ("8002210111110", "10002713", "EXTRAVERGINE", "SAGRA EXV BOT W12x1L CLASS IT", "Ex.v. Sagra Classico lt.1", 1.0, 10.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210133440", "10003255", "EXTRAVERGINE", "SAGRA EXV 100%R-PET V12x750ML IT", "Ex.v. Sagra lt.0,75 PET", 0.75, 7.50, "Pet.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210115088", "10002716", "EXTRAVERGINE", "SAGRA GRAND EXV BOT W12x1L", "Ex.v. Sagra Grandulivo lt.1", 1.0, 10.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210127562", "10002719", "EXTRAVERGINE", "SAGRA T.VIVE EXV BOT W 12x1L", "Ex.v. Sagra Terre Vive lt.1", 1.0, 10.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210119543", "10000536", "EXTRAVERGINE", "SAGRA PROF. EXV PET C2x5L IT", "Ex.v. Sagra Prof Lt.5", 5.0, 50.00, "Pet lt 5", 2, 17, 4, 68, 14, 9),
        ("8002210112827", "10002714", "EXTRAVERGINE", "SAGRA EXV 100%I BSA BOT W12x1L IT", "Ex.v. Sagra Bassa Acidità 100% ITA lt.1", 1.0, 15.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210127425", "10002715", "EXTRAVERGINE", "SAGRA EXV 100%I BOT W 12x1L", "Ex.v. Sagra 100% Italiano lt.1", 1.0, 15.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210128286", "10002720", "EXTRAVERGINE", "SAGRA EXV 100%I BIO BOT V12x1L IT", "Ex.v. Sagra Biologico 100% ITA lt.1", 1.0, 15.00, "Vetro lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210128248", "10002747", "EXTRAVERGINE", "SAGRA EXV BOT W12x750ML CLASS IT", "Ex.v. Sagra Classico lt.0,75", 0.75, 7.50, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210121997", "10003315", "EXTRAVERGINE", "SAGRA GRAND EXV BOT W12x750ML  IT", "Ex.v. Sagra Grandulivo 0,75", 0.75, 7.50, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210127197", "10003316", "EXTRAVERGINE", "SAGRA EXV 100%I BSA BOT W12x 750ML IT", "Ex.v. Sagra Bassa Acidità 100% ITA 0,75", 0.75, 11.25, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210133792", "10003317", "EXTRAVERGINE", "SAGRA EXV 100% I BOT W 12x750ML IT", "Ex.v. Sagra 100% Italiano 0,75", 0.75, 11.25, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210131815", "10003319", "EXTRAVERGINE", "SAGRA EXV 100%I BIO BOT W12x750ML IT", "Ex.v. Sagra Biologico 100% ITA  0,75", 0.75, 11.25, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210130814", "60000444", "EXTRAVERGINE", "SAGRA EXV SPRAY C6x200ML ALLUMINIO IT", "Ex.v. Sagra Spray ml.200", 0.2, 2.00, "Spray Lt 0,20", 6, 49, 6, 294, 14, 9),
        ("8002210124387", "10003061", "EXTRAVERGINE", "SAGRA PROF EXV PET T6x2L IT", "Ex.v. Sagra Prof lt.2", 2.0, 20.00, "Pet.Lt 2", 6, 13, 4, 52, 14, 9),
        ("8002210131620", "10002724", "EXTRAVERGINE", "FBERIO EXV BOT W12x1L CLASS IT", "Ex.v. Filippo Berio Classico lt.1", 1.0, 12.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210131644", "10002725", "EXTRAVERGINE", "FBERIO EXV BOT W12x1L BSA IT", "Ex.v. Filippo Berio Bassa Acidità lt.1", 1.0, 17.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210131705", "10002726", "EXTRAVERGINE", "FBERIO EXV 100%I BOT W12x1L IT", "Ex.v. Filippo Berio 100% Italiano lt.1", 1.0, 18.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210131767", "10002765", "EXTRAVERGINE", "FBERIO EXV BOT W12x750ML CLASS IT", "Ex.v. Filippo Berio Classico lt.0,75", 0.75, 12.00, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210131668", "10002746", "EXTRAVERGINE", "FBERIO EXV BSA BOT W12x750ML IT", "Ex.v. Filippo Berio Bassa Acidità lt.0,75", 0.75, 17.00, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210131804", "10002768", "EXTRAVERGINE", "FBERIO EXV 100%I BOT W12x750ML IT", "Ex.v. Filippo Berio 100% Italiano lt.0,75", 0.75, 18.00, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210133013", "10003200", "EXTRAVERGINE", "FB R.O. EXV BIO 100%IT MB BOT W12X750 IT", "Ex.v. Filippo Berio Riserva Oro lt.0,75", 0.75, 19.00, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210121461", "60000544", "EXTRAVERGINE", "EX.V. BUSTINA 10mlx250 FILIPPO BERIO ITA", "Ex.v. Filippo Berio Bustina ml.10", 0.01, 0.12, "bust lt 0,01", 250, 20, 5, 100, 14, 9),
        ("8002210126572", "10003240", "OLIVA", "SAGRA OOL PUR R-PET V12X750ML CLASS IT", "Oliva Sagra RPET lt.0,75 PET", 0.75, 8.00, "Pet.Lt 0,75", 12, 12, 5, 60, 18, 12),
        ("8002210001305", "10002717", "OLIVA", "SAGRA OOL BOT W12x1L CLASS", "Oliva Sagra lt.1", 1.0, 8.00, "Bott.Lt 1", 12, 8, 5, 40, 18, 12),
        ("8002210128453", "10002718", "OLIVA", "SAGRA GRAND OOL BOT W12x1L", "Oliva Sagra Grandulivo lt.1", 1.0, 8.00, "Bott.Lt 1", 12, 8, 5, 40, 18, 12),
        ("8002210126176", "10003288", "OLIVA", "SAGRA OOL PUR R-PET T6X1.5L IT", "Oliva Sagra lt.1,5", 1.5, 12.00, "Pet.Lt 1,5", 6, 16, 4, 64, 18, 12),
        ("8002210119567", "10000537", "OLIVA", "SAGRA PROF. OOL PUR PET C2x5L IT", "Oliva Sagra Prof Lt.5", 5.0, 40.00, "Pet.Lt 5", 2, 17, 4, 68, 18, 12),
        ("8002210132436", "10002965", "OLIVA", "FBERIO OOL PUR BOT V6X500ML IT", "Oliva Filippo Berio lt.0,50", 0.5, 4.11, "Bott.Lt 0,5", 6, 30, 6, 180, 18, 12),
        ("8002210131729", "10002727", "OLIVA", "FBERIO OOL PUR BOT W12x1L IT", "Oliva Filippo Berio lt.1", 1.0, 7.75, "Bott.Lt 1", 12, 8, 5, 40, 18, 12),
        ("8002210131781", "10002766", "OLIVA", "FBERIO OOL PUR BOT W12x750ML IT", "Oliva Filippo Berio lt.0,75", 0.75, 5.97, "Bott.Lt 0,75", 12, 12, 5, 60, 18, 12),
        ("8002210122307", "10000922", "OLIVA", "FBERIO OOL PUR LAT V8x1L IT", "Oliva Filippo Berio Latta lt.1", 1.0, 8.10, "Latta lt 1", 8, 12, 5, 60, 18, 12),
        ("8002210111486", "10003307", "SEMI", "SAGRA SEM MAIS PET V12x1L IT", "Mais Sagra lt.1", 1.0, 2.00, "Pet.Lt 1", 12, 12, 5, 60, 18, 12),
        ("8002210127067", "10003286", "SEMI", "SAGRA SEM MAIS PET T6x1.5L IT", "Mais Sagrì lt.1,5", 1.5, 3.00, "Pet.Lt 1,5", 6, 16, 4, 64, 18, 12),
        ("8002210112889", "10003089", "SEMI", "SAGRA SEM MAIS PET T6x2L IT", "Mais Sagra lt.2", 2.0, 4.00, "Pet.Lt 2", 6, 13, 4, 52, 18, 12),
        ("8002210000551", "10003311", "SEMI", "SAGRA SEM ARACHIDE PET V12x1L IT", "Arachide Sagra lt.1", 1.0, 3.00, "Pet.Lt 1", 12, 12, 5, 60, 18, 12),
        ("8002210126916", "10003284", "SEMI", "SAGRI SEM ARACHIDE PET T6x1.5L IT", "Arachide Sagrì lt.1,5", 1.5, 4.50, "Pet.Lt 1,5", 6, 16, 4, 64, 18, 12),
        ("8002210112865", "10003086", "SEMI", "SAGRA SEM ARACHIDE PET T6x2L IT", "Arachide Sagra lt.2", 2.0, 6.00, "Pet.Lt 2", 6, 13, 4, 52, 18, 12),
        ("8002210116160", "10000326", "SEMI", "SAGRA PROF SEM ARACHIDE PET C2x5L IT", "Arachide Sagra Prof. Lt.5", 5.0, 15.00, "Pet lt 5", 2, 17, 4, 68, 18, 12),
        ("8002210111905", "10003310", "SEMI", "SAGRA SEM GIRAS PET V12x1L IT", "Girasole Sagra lt.1", 1.0, 2.20, "Pet.Lt 1", 12, 12, 5, 60, 18, 12),
        ("8002210126817", "10003287", "SEMI", "SAGRI SEM GIRAS PET T6x1.5L IT", "Girasole Sagrì lt.1,5", 1.5, 3.30, "Pet.Lt 1,5", 6, 16, 4, 64, 18, 12),
        ("8002210113107", "10003087", "SEMI", "SAGRA SEM GIRAS PET T6x2L IT", "Girasole Sagra lt.2", 2.0, 4.40, "Pet.Lt 2", 6, 13, 4, 52, 18, 12),
        ("8002210115453", "10003062", "SEMI", "SAGRA PROF SEM GIRAS PET C2x5L IT", "Girasole Sagra Prof Lt.5", 5.0, 11.00, "Pet lt 5", 2, 17, 4, 68, 18, 12),
        ("8002210111295", "10002933", "SEMI", "SAGRA FRIMX SEM FRITT PET V12x1L NOP IT", "Frimax Sagra lt.1", 1.0, 2.25, "Pet Lt 1", 12, 12, 5, 60, 18, 12),
        ("8002210126893", "10003285", "SEMI", "SAGRI SEM FRITT PET T6x1.5L IT", "Frimax Sagrì lt.1,5", 1.5, 3.38, "Pet.Lt 1,5", 6, 16, 4, 64, 18, 12),
        ("8002210112940", "10003085", "SEMI", "SAGRA FRIMX SEM FRITT PET T6x2L NOP IT", "Frimax Sagra lt.2", 2.0, 4.50, "Pet Lt 2", 6, 13, 4, 52, 18, 12),
        ("8002210115484", "10002644", "SEMI", "SAGRA FRIMX SEM FRITT PET C2x5L NOP IT", "Frimax Sagra lt.5", 5.0, 11.25, "Pet Lt 5", 2, 17, 4, 68, 18, 12),
        ("8002210134140", "10003327", "SEMI", "GRAZIA SEM GIRAS LAT 1x20L IT", "Frimax Spray ml.200", 0.2, 0.45, "Spray Lt 0,20", 6, 49, 6, 294, 18, 12),
        ("8002210127401", "10003309", "SEMI", "SAGRA SEM GIRAS AO PET V12x1L IT", "Girasole Alto Oleico Sagra lt.1", 1.0, 2.80, "Pet.Lt 1", 12, 12, 5, 60, 18, 12),
        ("8002210126336", "10003063", "SEMI", "SAGRA PROF SEM GIRAS AO PET C2x5L IT", "Girasole Alto Oleico Sagra Prof lt.5", 5.0, 14.00, "Pet lt 5", 2, 17, 4, 68, 18, 12),
        ("8002210129290", "10003312", "SEMI", "SAGRA SEM VINACC PET V12x1L IT", "Vinacciolo Sagra lt.1", 1.0, 5.00, "Pet.Lt 1", 12, 12, 5, 60, 18, 12),
        ("8002210130289", "10003082", "EXTRAVERGINE", "FBERIO EXV CLASS MB BOT V6x250ML IT", "Ex.v. F.Berio Anti Rab Classico lt.0,25", 0.25, 2.50, "Vetro lt 0,25", 6, 49, 5, 245, 14, 9),
        ("8002210130210", "10003081", "EXTRAVERGINE", "FBERIO EXV 100%I MB BOT V6x250ML IT", "Ex.v. F.Berio Anti Rab 100% ITA lt.0,25", 0.25, 3.00, "Vetro lt 0,25", 6, 49, 5, 245, 14, 9),
        ("8002210130340", "10003091", "EXTRAVERGINE", "FBERIO EXV CLASS MB BOT V6x500ML IT", "Ex.v. F.Berio Anti Rab Classico lt.0,50", 0.5, 4.30, "Vetro lt 0,50", 6, 31, 5, 155, 14, 9),
        ("8002210130302", "10003079", "EXTRAVERGINE", "FBERIO EXV 100%I MB BOT V6x500ML IT", "Ex.v. F.Berio Anti Rab 100% ITA lt.0,50", 0.5, 4.80, "Vetro lt 0,50", 6, 31, 5, 155, 14, 9),
        ("8002210132573", "10003072", "EXTRAVERGINE", "FBERIO EXV BOT V6x500ML TOSC IT", "Ex.v. F.Berio Toscano lt.0,50", 0.5, 10.00, "Vetro lt 0,50", 6, 31, 5, 155, 18, 12),
        ("8002210130234", "60000591", "EXTRAVERGINE", "FBERIO EXV DRES BOT V6x250ML PEP TE IT", "Ex.v. F.Berio Peperoncino lt.0,25", 0.25, 3.50, "Vetro lt 0,25", 6, 49, 5, 245, 24, 16),
        ("8002210130791", "60000590", "ACETO", "FBERIO ACE BALS BOT V6x250ML IT", "Aceto Balsamico F.Berio lt.0,25", 0.25, 2.00, "Vetro lt 0,25", 6, 48, 6, 288, 61, 41),
        ("8002210130197", "60000589", "ACETO", "FBERIO ACE BALS BOT V6x500ML IT", "Aceto Balsamico F.Berio lt.0,50", 0.5, 2.10, "Vetro lt 0,50", 6, 31, 5, 155, 61, 41)
    ]
    
    for p in prodotti_salov:
        cursor.execute("""
        INSERT OR REPLACE INTO anagrafica_master (
            ean, codice_sap, tipo_olio, descrizione_sap, descrizione_commerciale, formato_lt, confezione,
            pezzi_cartone, cartoni_strato, strati_pallet, cartoni_pallet, conservazione_mesi, shelf_life_mesi
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (p[0], p[1], p[2], p[3], p[4], p[5], p[7], p[8], p[9], p[10], p[11], p[12], p[13]))
        
        cursor.execute("""
        INSERT OR REPLACE INTO guardrail_aziendali (ean, min_net_net_g) VALUES (?, ?)
        """, (p[0], p[6]))
        
    clienti_demo = [
        ("COOP ITALIA", "COOP ITALIA SOTTOGRUPPO", "ALLEANZA 3.0"),
        ("CONAD", "CONAD SOTTOGRUPPO", "CONAD ADRIATICO"),
        ("ESSELUNGA GRUPPO", "ESSELUNGA SOTTOGRUPPO", "ESSELUNGA"),
        ("SELEX GRUPPO", "SELEX SOTTOGRUPPO", "SELEX "),
        ("PAM GRUPPO", "PAM SOTTOGRUPPO", "PAM"),
        ("CRAI GRUPPO", "CRAI SOTTOGRUPPO", "CRAI TIRRENO")
    ]
    for c in clienti_demo:
        cursor.execute("INSERT OR IGNORE INTO clienti (gruppo_macro, sottogruppo, associato_insegna) VALUES (?, ?, ?)", c)
        
    fallback_data = [
        ('COOP ITALIA', '', '', 'GRUPPO', '', None, 20.0, 30.0, None, None, None, None, None, None, 1.5, 1.0, 14.0, 8.0, None, None, None),
        ('COOP ITALIA', 'COOP ITALIA SOTTOGRUPPO', 'ALLEANZA 3.0', 'REFERENZA', '8002210131620', 66.00, None, None, None, None, None, None, 12.0, 5.0, None, None, None, None, None, None, None),
        ('COOP ITALIA', 'COOP ITALIA SOTTOGRUPPO', 'ALLEANZA 3.0', 'REFERENZA', '8002210111110', 60.80, None, None, None, None, None, None, 15.0, 0.0, None, None, None, None, None, None, None),
        ('COOP ITALIA', 'COOP ITALIA SOTTOGRUPPO', 'ALLEANZA 3.0', 'REFERENZA', '8002210001305', 43.20, None, None, None, None, None, None, 12.0, 0.0, None, None, None, None, None, None, None),

        ('ESSELUNGA GRUPPO', '', '', 'GRUPPO', '', None, 35.0, 15.0, None, None, None, None, None, None, 1.2, 1.0, 12.0, 5.0, None, None, None),
        ('ESSELUNGA GRUPPO', 'ESSELUNGA SOTTOGRUPPO', 'ESSELUNGA', 'REFERENZA', '8002210131620', 40.00, None, None, None, None, None, None, 10.0, 7.0, None, None, None, None, None, None, None),
        ('ESSELUNGA GRUPPO', 'ESSELUNGA SOTTOGRUPPO', 'ESSELUNGA', 'REFERENZA', '8002210111110', 38.00, None, None, None, None, None, None, 11.0, 0.0, None, None, None, None, None, None, None),
        ('ESSELUNGA GRUPPO', 'ESSELUNGA SOTTOGRUPPO', 'ESSELUNGA', 'REFERENZA', '8002210001305', 24.00, None, None, None, None, None, None, 13.0, 0.0, None, None, None, None, None, None, None),

        ('CONAD', '', '', 'GRUPPO', '', None, 17.0, 18.0, None, None, None, None, None, None, 1.5, 1.0, 9.0, 11.0, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', 'CONAD ADRIATICO', 'REFERENZA', '8002210131620', 50.00, None, None, None, None, None, None, 12.0, 9.0, None, None, None, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', 'CONAD ADRIATICO', 'REFERENZA', '8002210111110', 44.00, None, None, None, None, None, None, 11.0, 4.0, None, None, None, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', 'CONAD ADRIATICO', 'REFERENZA', '8002210001305', 30.00, None, None, None, None, None, None, 10.0, 4.0, None, None, None, None, None, None, None),

        ('SELEX GRUPPO', '', '', 'GRUPPO', '', None, 17.0, 18.0, None, None, None, None, None, None, 1.5, 1.0, 9.0, 11.0, None, None, None),
        ('SELEX GRUPPO', 'SELEX SOTTOGRUPPO', 'SELEX ', 'REFERENZA', '8002210131620', 50.00, None, None, None, None, None, None, 12.0, 9.0, None, None, None, None, None, None, None),
        ('SELEX GRUPPO', 'SELEX SOTTOGRUPPO', 'SELEX ', 'REFERENZA', '8002210111110', 44.00, None, None, None, None, None, None, 11.0, 4.0, None, None, None, None, None, None, None),
        ('SELEX GRUPPO', 'SELEX SOTTOGRUPPO', 'SELEX ', 'REFERENZA', '8002210001305', 30.00, None, None, None, None, None, None, 10.0, 4.0, None, None, None, None, None, None, None),
        
        ('PAM GRUPPO', '', '', 'GRUPPO', '', None, 15.0, 20.0, None, None, None, None, None, None, 1.4, 1.0, 11.0, 6.0, None, None, None),
        ('PAM GRUPPO', 'PAM SOTTOGRUPPO', 'PAM', 'REFERENZA', '8002210131620', 52.00, None, None, None, None, None, None, 14.0, 6.0, None, None, None, None, None, None, None),
        ('PAM GRUPPO', 'PAM SOTTOGRUPPO', 'PAM', 'REFERENZA', '8002210111110', 48.00, None, None, None, None, None, None, 13.0, 3.0, None, None, None, None, None, None, None),
        ('PAM GRUPPO', 'PAM SOTTOGRUPPO', 'PAM', 'REFERENZA', '8002210001305', 32.00, None, None, None, None, None, None, 9.0, 3.0, None, None, None, None, None, None, None),

        ('CRAI GRUPPO', '', '', 'GRUPPO', '', None, 12.0, 25.0, None, None, None, None, None, None, 2.0, 1.0, 7.0, 12.0, None, None, None),
        ('CRAI GRUPPO', 'CRAI SOTTOGRUPPO', 'CRAI TIRRENO', 'REFERENZA', '8002210131620', 56.00, None, None, None, None, None, None, 15.0, 8.0, None, None, None, None, None, None, None),
        ('CRAI GRUPPO', 'CRAI SOTTOGRUPPO', 'CRAI TIRRENO', 'REFERENZA', '8002210111110', 50.00, None, None, None, None, None, None, 12.0, 5.0, None, None, None, None, None, None, None),
        ('CRAI GRUPPO', 'CRAI SOTTOGRUPPO', 'CRAI TIRRENO', 'REFERENZA', '8002210001305', 35.00, None, None, None, None, None, None, 11.0, 5.0, None, None, None, None, None, None, None)
    ]
    
    cursor.executemany("""
    INSERT OR REPLACE INTO accordi_commerciali (
        gruppo_macro, sottogruppo, associato_insegna, livello, chiave_livello, listino_r,
        sconto_1, sconto_2, sconto_3, sconto_4, sconto_5,
        sconto_6, sconto_7, sconto_y, sconto_carico, sconto_pagamento,
        voce_contratto_1, voce_contratto_2, voce_contratto_3, voce_contratto_4, voce_contratto_5
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, fallback_data)
    conn.commit()
    conn.close()

init_db()

menu = st.sidebar.radio("SELEZIONA SCHEDA", [
    "Simulatore Offerte", 
    "Storico Promozioni", 
    "Dati Anagrafici (Logistica)", 
    "Back-Office (Contratti)", 
    "Report Sintetico", 
    "Guida Operativa"
])


# ==========================================
# SCHEDA 1: SIMULATORE
# ==========================================
if menu == "Simulatore Offerte":
    st.title("Commerciale Salov - Simulatore")
    conn = sqlite3.connect(DB_FILE)
    
    st.sidebar.header("Parametri Negoziazione")
    
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT gruppo_macro FROM clienti WHERE attivo=1 ORDER BY gruppo_macro")
    gruppi = [r[0] for r in cursor.fetchall()]
    
    if not gruppi:
        st.warning("ATTENZIONE: Nessun cliente caricato. Sblocca il sistema caricando i dati dal Back-Office.")
    else:
        gruppo_sel = st.sidebar.selectbox("1. Gruppo GDO", gruppi, help="Seleziona la centrale d'acquisto.")
        
        cursor.execute("SELECT DISTINCT sottogruppo FROM clienti WHERE gruppo_macro=? AND attivo=1 ORDER BY sottogruppo", (gruppo_sel,))
        sottogruppi = [r[0] for r in cursor.fetchall()]
        sottogruppo_sel = st.sidebar.selectbox("2. Sottogruppo GDO", sottogruppi, help="Seleziona il sottogruppo di canale.")
        
        cursor.execute("SELECT DISTINCT associato_insegna FROM clienti WHERE gruppo_macro=? AND sottogruppo=? AND attivo=1 ORDER BY associato_insegna", (gruppo_sel, sottogruppo_sel))
        associati = [r[0] for r in cursor.fetchall()]
        associato_sel = st.sidebar.selectbox("3. Insegna Locale / Associato", associati, help="Seleziona l'associato locale.")

        cursor.execute("""
            SELECT a.ean, a.descrizione_commerciale, a.tipo_olio, COALESCE(g.min_net_net_g, 0.0), a.codice_sap, a.formato_lt,
                   a.pezzi_cartone, a.cartoni_strato, a.strati_pallet, a.cartoni_pallet
            FROM anagrafica_master a
            LEFT JOIN guardrail_aziendali g ON a.ean = g.ean
        """)
        prodotti = cursor.fetchall()
        prodotti_dict = {f"{p[1]} [EAN: {p[0]}]": (p[0], p[2], p[3], p[4], p[5], p[6], p[7], p[8], p[9]) for p in prodotti}
        prodotto_scelto = st.sidebar.selectbox("4. Referenza Salov", list(prodotti_dict.keys()), help="Seleziona la referenza.")
        
        ean, tipo_olio, min_net_net_g, codice_sap, formato_lt, pezzi_cartone, cartoni_strato, strati_pallet, cartoni_pallet = prodotti_dict[prodotto_scelto]

        contract = HierarchyResolver.resolve(conn, gruppo_sel, sottogruppo_sel, associato_sel, ean, tipo_olio)

        st.sidebar.markdown("---")
        st.sidebar.subheader("Verità Contrattuale")
        if contract.listino_r is None:
            st.error("ATTENZIONE: PRODOTTO FUORI ASSORTIMENTO PER QUESTO CLIENTE")
            st.stop()
        else:
            st.sidebar.info(f"Listino Base (R): {contract.listino_r:.2f} Euro")
            st.sidebar.text(f"Livello Risolto: {contract.livello_risolto}")

        st.markdown("### Scegli Metodologia di Negoziazione")
        metodo_lavoro = st.radio(
            "Seleziona l'approccio negoziale:",
            ["A. Partenza da Prezzo Target (Calcolo automatico Sconto Promo)", "B. Tentativi Spot Manuali (Immissione Sconto Promo libera)"],
            horizontal=True
        )

        st.markdown("---")
        
        if "A. Partenza" in metodo_lavoro:
            st.markdown("### Definizione Obiettivo Economico")
            target_container = st.container(border=True)
            with target_container:
                col_t1, col_t2 = st.columns([2, 1])
                with col_t1:
                    target_net_net = st.number_input(
                        "PREZZO TARGET NET NET DESIDERATO (Euro/Pz)", 
                        min_value=0.0, 
                        value=float(min_net_net_g), 
                        step=0.10,
                        help="Fissa il ricavo reale netto a bottiglia (AM) che desideri ottenere. Di default è la soglia minima di sicurezza (G)."
                    )
                with col_t2:
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.info("Modo Target Attivo")
        else:
            target_net_net = 0.0

        st.markdown("---")
        st.subheader("Manovre e Leve Sconti")
        
        col_y1, col_y2 = st.columns(2)
        with col_y1:
            sconto_y = st.number_input(
                "Sconto Continuativo Y (%)", 
                min_value=0.0, max_value=100.0, 
                value=float(contract.sconto_y), step=0.5
            )
            st.markdown(
                f"<div class='warning-box'>ATTENZIONE - UNO SCONTO CONTINUATIVO PUO' DERIVARE DA UN ACCORDO LOCALE - valore attuale: {contract.sconto_y:.2f}%<br>"
                f"<span style='font-size:0.8em; font-weight:normal;'>La modifica in corsa potrebbe violare tale accordo che va quindi ridiscusso prima di confermare la promozione.</span></div>", 
                unsafe_allow_html=True
            )
            
        st.markdown("<br>", unsafe_allow_html=True)
        
        col_l1, col_l2 = st.columns(2)
        
        with col_l1:
            if "A. Partenza" in metodo_lavoro:
                aa_box = st.container(border=True)
                with aa_box:
                    st.markdown("<h4 style='color: #1A3E2F; margin-bottom: 5px;'>Leva Promozionale Diretta</h4>", unsafe_allow_html=True)
                    st.markdown("<span style='font-size: 0.9em; color: #4B5563;'>In Modalità Target lo Sconto Z è automatico. Usa questo campo per forzare un taglio prezzo unitario in fattura.</span>", unsafe_allow_html=True)
                    st.markdown("<br>", unsafe_allow_html=True)
                    sconto_aa = st.number_input(
                        "Sconto Unitario in fattura (Euro/Pz) [AA]", 
                        min_value=0.0, value=0.0, step=0.05
                    )
            else:
                st.markdown("**Leve Promozionali**")
                sconto_z_input = st.number_input("Sconto Promozionale (%) [Z] (Manuale)", min_value=0.0, max_value=100.0, value=0.0, step=0.5)
                sconto_z = Decimal(f"{sconto_z_input:.5f}")
                
                sconto_aa = st.number_input(
                    "Sconto Unitario in fattura (Euro/Pz) [AA]", 
                    min_value=0.0, value=0.0, step=0.05
                )

        if "A. Partenza" in metodo_lavoro:
            target_dec = Decimal(f"{target_net_net:.5f}")
            temp_input = PricingInput(
                listino_r=contract.listino_r,
                sconto_1=contract.sconto_1, sconto_2=contract.sconto_2, sconto_3=contract.sconto_3,
                sconto_4=contract.sconto_4, sconto_5=contract.sconto_5, sconto_6=contract.sconto_6, sconto_7=contract.sconto_7,
                sconto_y=Decimal(f"{sconto_y:.5f}"), sconto_z=Decimal("0.00"), sconto_aa=Decimal(f"{sconto_aa:.5f}"),
                sconto_carico=contract.sconto_carico, sconto_pagamento=contract.sconto_pagamento,
                voce_i=contract.voce_i, voce_ii=contract.voce_ii, voce_iii=contract.voce_iii, voce_iv=contract.voce_iv, voce_v=contract.voce_v,
                min_net_net_g=Decimal(str(min_net_net_g))
            )
            sconto_z = PricingEngine.calculate_inverse(target_dec, temp_input, "Z")

        with col_l2:
            if "A. Partenza" in metodo_lavoro:
                st.markdown("**Analisi Limiti Promozionali**")
                temp_input_max_z = PricingInput(
                    listino_r=contract.listino_r,
                    sconto_1=contract.sconto_1, sconto_2=contract.sconto_2, sconto_3=contract.sconto_3,
                    sconto_4=contract.sconto_4, sconto_5=contract.sconto_5, sconto_6=contract.sconto_6, sconto_7=contract.sconto_7,
                    sconto_y=Decimal(f"{sconto_y:.5f}"), sconto_z=Decimal("0.00"), sconto_aa=Decimal(f"{sconto_aa:.5f}"),
                    sconto_carico=contract.sconto_carico, sconto_pagamento=contract.sconto_pagamento,
                    voce_i=contract.voce_i, voce_ii=contract.voce_ii, voce_iii=contract.voce_iii, voce_iv=contract.voce_iv, voce_v=contract.voce_v,
                    min_net_net_g=Decimal(str(min_net_net_g))
                )
                z_max_consentito = PricingEngine.calculate_inverse(Decimal(str(min_net_net_g)), temp_input_max_z, "Z")
                st.number_input("Sconto Promo MAX Consentito [Z]", value=float(z_max_consentito), disabled=True, format="%.2f", help="Il massimo Sconto Z che puoi inserire (a parità di AA) prima di andare in blocco.")
                
            else:
                st.markdown("**Analisi Limiti Promozionali**")
                temp_input_max_z = PricingInput(
                    listino_r=contract.listino_r,
                    sconto_1=contract.sconto_1, sconto_2=contract.sconto_2, sconto_3=contract.sconto_3,
                    sconto_4=contract.sconto_4, sconto_5=contract.sconto_5, sconto_6=contract.sconto_6, sconto_7=contract.sconto_7,
                    sconto_y=Decimal(f"{sconto_y:.5f}"), sconto_z=Decimal("0.00"), sconto_aa=Decimal(f"{sconto_aa:.5f}"),
                    sconto_carico=contract.sconto_carico, sconto_pagamento=contract.sconto_pagamento,
                    voce_i=contract.voce_i, voce_ii=contract.voce_ii, voce_iii=contract.voce_iii, voce_iv=contract.voce_iv, voce_v=contract.voce_v,
                    min_net_net_g=Decimal(str(min_net_net_g))
                )
                z_max_consentito = PricingEngine.calculate_inverse(Decimal(str(min_net_net_g)), temp_input_max_z, "Z")
                st.number_input("Sconto Promo MAX Consentito [Z]", value=float(z_max_consentito), disabled=True, format="%.2f")
                
                temp_input_max_aa = PricingInput(
                    listino_r=contract.listino_r,
                    sconto_1=contract.sconto_1, sconto_2=contract.sconto_2, sconto_3=contract.sconto_3,
                    sconto_4=contract.sconto_4, sconto_5=contract.sconto_5, sconto_6=contract.sconto_6, sconto_7=contract.sconto_7,
                    sconto_y=Decimal(f"{sconto_y:.5f}"), sconto_z=Decimal(f"{sconto_z:.5f}"), sconto_aa=Decimal("0.00"),
                    sconto_carico=contract.sconto_carico, sconto_pagamento=contract.sconto_pagamento,
                    voce_i=contract.voce_i, voce_ii=contract.voce_ii, voce_iii=contract.voce_iii, voce_iv=contract.voce_iv, voce_v=contract.voce_v,
                    min_net_net_g=Decimal(str(min_net_net_g))
                )
                aa_max_consentito = PricingEngine.calculate_inverse(Decimal(str(min_net_net_g)), temp_input_max_aa, "AA")
                st.number_input("Sconto Unitario MAX Consentito [AA]", value=float(aa_max_consentito), disabled=True, format="%.2f")

        engine_input = PricingInput(
            listino_r=contract.listino_r,
            sconto_1=contract.sconto_1, sconto_2=contract.sconto_2, sconto_3=contract.sconto_3,
            sconto_4=contract.sconto_4, sconto_5=contract.sconto_5, sconto_6=contract.sconto_6, sconto_7=contract.sconto_7,
            sconto_y=Decimal(f"{sconto_y:.5f}"), sconto_z=sconto_z, sconto_aa=Decimal(f"{sconto_aa:.5f}"),
            sconto_carico=contract.sconto_carico, sconto_pagamento=contract.sconto_pagamento,
            voce_i=contract.voce_i, voce_ii=contract.voce_ii, voce_iii=contract.voce_iii, voce_iv=contract.voce_iv, voce_v=contract.voce_v,
            min_net_net_g=Decimal(str(min_net_net_g))
        )
        result = PricingEngine.calculate(engine_input)

        st.markdown("---")
        
        col_c1, col_c2 = st.columns(2)
        
        with col_c1:
            with st.expander("Verifica Margine e Stato (Contrattuale)", expanded=True):
                st.metric("PREZZO NET NET RISULTANTE (AM)", f"{result.net_net_finale:.2f} Euro")
                st.metric("SOGLIA MINIMA NET NET (G)", f"{min_net_net_g:.2f} Euro")
                if result.guardrail_ok:
                    st.success(f"VERDE (APPROVATO) - Margine sicuro. Delta: +{result.delta_vs_min:.2f} Euro")
                else:
                    st.error(f"BLOCCATO! SI PERDE SOLDI !!! - Sotto soglia di {abs(result.delta_vs_min):.2f} Euro")
        
        with col_c2:
            with st.expander("Finestra Temporale Promo", expanded=True):
                col_d1, col_d2 = st.columns(2)
                with col_d1:
                    sell_in_dal = st.date_input("Inizio Sell-In", date.today(), key="si_dal")
                    sell_in_al = st.date_input("Fine Sell-In", date.today(), key="si_al")
                with col_d2:
                    sell_out_dal = st.date_input("Inizio Sell-Out", date.today(), key="so_dal")
                    sell_out_al = st.date_input("Fine Sell-Out", date.today(), key="so_al")

        st.markdown("---")

        st.markdown("### Contributi Promozionali Extra (Volantino / Sell-Out)")
        st.markdown("<span style='font-size: 0.9em; color: #4B5563;'>Inserisci eventuali costi extra richiesti dalla GDO per l'operazione. Se non inserisci i volumi, il costo fisso verrà registrato ma non impatterà il calcolo unitario.</span>", unsafe_allow_html=True)
        
        box_sellout = st.container(border=True)
        with box_sellout:
            col_v1, col_v2, col_v3 = st.columns(3)
            with col_v1:
                volumi_stimati = st.number_input("Volumi Stimati (Pezzi)", min_value=0, value=0, step=100, help="Numero di bottiglie previste per la promo.")
            with col_v2:
                contributo_fisso = st.number_input("Contributo Fisso Totale (€)", min_value=0.0, value=0.0, step=50.0, help="Es. Costo testata gondola, inserimento volantino.")
            with col_v3:
                contributo_pezzo = st.number_input("Contributo a Pezzo (€/Pz)", min_value=0.0, value=0.0, step=0.05, help="Es. Contributo di 0.10€ per ogni pezzo venduto in cassa.")

            costo_totale_extra = contributo_fisso + (contributo_pezzo * volumi_stimati)
            impatto_unitario_extra = Decimal("0.00")
            net_net_post_promo = result.net_net_finale
            mostra_impatto = False
            
            if (volumi_stimati > 0 and contributo_fisso > 0) or (contributo_pezzo > 0):
                mostra_impatto = True
                if volumi_stimati > 0:
                    impatto_unitario_extra = Decimal(str(contributo_pezzo)) + (Decimal(str(contributo_fisso)) / Decimal(str(volumi_stimati)))
                else:
                    impatto_unitario_extra = Decimal(str(contributo_pezzo))
                
                net_net_post_promo = result.net_net_finale - impatto_unitario_extra
                
                st.markdown("---")
                st.warning(f"**Costo Promozionale Extra Totale:** {costo_totale_extra:.2f} €")
                st.error(f"**Impatto Unitario Extra:** -{impatto_unitario_extra:.3f} €/Pz ➔ **NET NET POST-VOLANTINO: {net_net_post_promo:.3f} €**")
            elif costo_totale_extra > 0:
                st.markdown("---")
                st.warning(f"**Costo Promozionale Extra Totale:** {costo_totale_extra:.2f} €")
                st.info("ℹ️ Volumi non inseriti: l'impatto unitario del contributo fisso non è calcolabile, ma il costo totale verrà registrato nello storico.")

        st.markdown("---")
        st.subheader("Tabella Sequenziale Estesa della Struttura di Costo")
        
        waterfall_data = [
            {"Fase Pricing": step.fase, "Valore Unitario": step.valore, "Dettaglio Operazione": step.descrizione}
            for step in result.steps
        ]
        
        if mostra_impatto:
            waterfall_data.append({
                "Fase Pricing": "Impatto Extra (Sell-Out)", 
                "Valore Unitario": net_net_post_promo.quantize(Decimal("0.001")), 
                "Dettaglio Operazione": f"-{impatto_unitario_extra.quantize(Decimal('0.001'))} €/Pz"
            })
            
        df_waterfall = pd.DataFrame(waterfall_data)
        st.dataframe(df_waterfall, use_container_width=True, hide_index=True)

        st.markdown("---")
        
        st.markdown("### Storicizzazione Promozione")
        box_salvataggio = st.container(border=True)
        with box_salvataggio:
            col_s1, col_s2, col_s3 = st.columns([1, 2, 1])
            with col_s1:
                stato_promo = st.radio("Stato della Promozione:", ["Proposta", "Confermata"])
            with col_s2:
                note_promo = st.text_area("Note opzionali (es. Riferimento Volantino, Accordi verbali):", height=68)
            with col_s3:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("SALVA PROMOZIONE NEL DATABASE", type="primary", use_container_width=True):
                    try:
                        c_save = conn.cursor()
                        c_save.execute("""
                            INSERT INTO storico_promo (
                                stato_promo, gruppo_macro, sottogruppo, associato_insegna,
                                ean, descrizione_commerciale, listino_r, sconto_y, sconto_z, sconto_aa,
                                net_net_am, volumi_stimati, contributo_fisso, contributo_pezzo, costo_totale_extra, note,
                                sell_in_dal, sell_in_al, sell_out_dal, sell_out_al, min_net_net_g, net_net_post_promo
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            stato_promo, gruppo_sel, sottogruppo_sel, associato_sel,
                            ean, prodotto_scelto.split(" [EAN:")[0],
                            float(contract.listino_r), float(sconto_y), float(sconto_z), float(sconto_aa),
                            float(result.net_net_finale),
                            volumi_stimati, contributo_fisso, contributo_pezzo, costo_totale_extra, note_promo,
                            sell_in_dal.strftime('%Y-%m-%d'), sell_in_al.strftime('%Y-%m-%d'), 
                            sell_out_dal.strftime('%Y-%m-%d'), sell_out_al.strftime('%Y-%m-%d'),
                            float(min_net_net_g), float(net_net_post_promo)
                        ))
                        conn.commit()
                        st.success(f"✅ Promozione salvata con successo come '{stato_promo}'!")
                    except Exception as e:
                        st.error(f"Errore durante il salvataggio: {e}")

        st.markdown("---")
        
        def genera_scheda_negoziale():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Proposta_Commerciale"
            ws.views.sheetView[0].showGridLines = True
            
            font_title = Font(name="Arial", size=15, bold=True, color="FFFFFF")
            font_section = Font(name="Arial", size=11, bold=True, color="000000")
            font_label = Font(name="Arial", size=10, bold=True)
            font_value = Font(name="Arial", size=10)
            fill_header = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
            fill_sub = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )
            
            ws.merge_cells('A1:D1')
            ws['A1'] = "SALOV S.p.A. - SCHEDA PROPOSTA COMMERCIALE"
            ws['A1'].font = font_title
            ws['A1'].fill = fill_header
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 40
            
            ws['A3'] = "ANAGRAFICA GDO"
            ws['A3'].font = font_section
            ws['A3'].fill = fill_sub
            ws.merge_cells('A3:D3')
            
            ws['A4'] = "Gruppo Macro:"
            ws['B4'] = gruppo_sel
            ws['A5'] = "Sottogruppo:"
            ws['B5'] = sottogruppo_sel
            ws['A6'] = "Insegna Locale / Associato:"
            ws['B6'] = associato_sel
            
            for r in range(4, 7):
                ws[f'A{r}'].font = font_label
                ws[f'B{r}'].font = font_value
            
            ws['A8'] = "DETTAGLIO REFERENZA"
            ws['A8'].font = font_section
            ws['A8'].fill = fill_sub
            ws.merge_cells('A8:D8')
            
            ws['A9'] = "Descrizione Articolo:"
            ws['B9'] = prodotto_scelto.split(" [EAN:")[0]
            ws['A10'] = "EAN:"
            ws['B10'] = ean
            ws['A11'] = "Codice SAP:"
            ws['B11'] = codice_sap
            ws['A12'] = "Formato:"
            ws['B12'] = f"{formato_lt} Litri"
            
            for r in range(9, 13):
                ws[f'A{r}'].font = font_label
                ws[f'B{r}'].font = font_value
                
            ws['A14'] = "DATI LOGISTICI E PALLETTIZZAZIONE"
            ws['A14'].font = font_section
            ws['A14'].fill = fill_sub
            ws.merge_cells('A14:D14')
            
            ws['A15'] = "Pezzi per Cartone:"
            ws['B15'] = pezzi_cartone if pezzi_cartone is not None else 0
            ws['A16'] = "Cartoni per Strato:"
            ws['B16'] = cartoni_strato if cartoni_strato is not None else 0
            ws['A17'] = "Strati per Pallet:"
            ws['B17'] = strati_pallet if strati_pallet is not None else 0
            ws['A18'] = "Cartoni per Pallet:"
            ws['B18'] = cartoni_pallet if cartoni_pallet is not None else 0
            ws['A19'] = "Pezzi Totali per Pallet:"
            ws['B19'] = (pezzi_cartone or 0) * (cartoni_pallet or 0)
            
            for r in range(15, 20):
                ws[f'A{r}'].font = font_label
                ws[f'B{r}'].font = font_value
                
            ws['A21'] = "FINESTRE TEMPORALI PROMO"
            ws['A21'].font = font_section
            ws['A21'].fill = fill_sub
            ws.merge_cells('A21:D21')
            
            ws['A22'] = "Periodo Sell-In:"
            ws['B22'] = f"Dal {sell_in_dal.strftime('%d/%m/%Y')} al {sell_in_al.strftime('%d/%m/%Y')}"
            ws['A23'] = "Periodo Sell-Out:"
            ws['B23'] = f"Dal {sell_out_dal.strftime('%d/%m/%Y')} al {sell_out_al.strftime('%d/%m/%Y')}"
            
            for r in range(22, 24):
                ws[f'A{r}'].font = font_label
                ws[f'B{r}'].font = font_value
            
            ws['A25'] = "CASCATA DI PRICING NEGOZIALE"
            ws['A25'].font = font_section
            ws['A25'].fill = fill_sub
            ws.merge_cells('A25:D25')
            
            ws['A26'] = "Elemento di Costo"
            ws['B26'] = "Valore"
            ws['C26'] = "Tipologia Operazione"
            for col in ['A', 'B', 'C']:
                ws[f'{col}26'].font = font_label
                
            row_idx = 27
            for step in result.steps:
                ws.cell(row=row_idx, column=1, value=step.fase).font = font_value
                ws.cell(row=row_idx, column=2, value=float(step.valore)).font = font_value
                ws.cell(row=row_idx, column=2).number_format = '#,##0.000 €'
                ws.cell(row=row_idx, column=3, value=step.descrizione).font = font_value
                row_idx += 1
                
            if mostra_impatto:
                ws.cell(row=row_idx, column=1, value="Impatto Extra (Sell-Out)").font = font_value
                ws.cell(row=row_idx, column=2, value=float(net_net_post_promo)).font = font_value
                ws.cell(row=row_idx, column=2).number_format = '#,##0.000 €'
                ws.cell(row=row_idx, column=3, value=f"-{float(impatto_unitario_extra):.3f} €/Pz").font = font_value
                row_idx += 1
                
            ws.cell(row=row_idx+1, column=1, value="SOGLIA MINIMA AM (G):").font = font_label
            ws.cell(row=row_idx+1, column=2, value=float(min_net_net_g)).font = font_value
            ws.cell(row=row_idx+1, column=2).number_format = '#,##0.00 €'
            
            ws.cell(row=row_idx+2, column=1, value="DELTA DI MARGINE VS SOGLIA:").font = font_label
            ws.cell(row=row_idx+2, column=2, value=float(result.delta_vs_min)).font = font_value
            ws.cell(row=row_idx+2, column=2).number_format = '#,##0.00 €'
            
            ws.cell(row=row_idx+3, column=1, value="STATO DEL MARGINE:").font = font_label
            stato_txt = "VERDE (APPROVATO)" if result.guardrail_ok else "ROSSO (SOTTO SOGLIA)"
            ws.cell(row=row_idx+3, column=2, value=stato_txt).font = font_label
            
            if costo_totale_extra > 0:
                row_idx += 5
                ws.cell(row=row_idx, column=1, value="CONTRIBUTI EXTRA (SELL-OUT)").font = font_section
                ws.cell(row=row_idx, column=1).fill = fill_sub
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
                
                ws.cell(row=row_idx+1, column=1, value="Volumi Stimati (Pz):").font = font_label
                ws.cell(row=row_idx+1, column=2, value=volumi_stimati).font = font_value
                
                ws.cell(row=row_idx+2, column=1, value="Costo Totale Extra:").font = font_label
                ws.cell(row=row_idx+2, column=2, value=costo_totale_extra).font = font_value
                ws.cell(row=row_idx+2, column=2).number_format = '#,##0.00 €'
                
                if volumi_stimati > 0:
                    ws.cell(row=row_idx+3, column=1, value="Net Net Post-Volantino:").font = font_label
                    ws.cell(row=row_idx+3, column=2, value=float(net_net_post_promo)).font = font_value
                    ws.cell(row=row_idx+3, column=2).number_format = '#,##0.000 €'
            
            ws.column_dimensions['A'].width = 32
            ws.column_dimensions['B'].width = 38
            ws.column_dimensions['C'].width = 45
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
                for cell in row:
                    cell.border = thin_border
            
            buffer = io.BytesIO()
            wb.save(buffer)
            return buffer.getvalue()

        proposta_excel = genera_scheda_negoziale()
        
        st.download_button(
            label="SCARICA PROPOSTA COMMERCIALE PER IL CLIENTE",
            data=proposta_excel,
            file_name=f"Proposta_{associato_sel}_{codice_sap}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ==========================================
# NUOVA SCHEDA: STORICO PROMOZIONI
# ==========================================
elif menu == "Storico Promozioni":
    st.title("Storico Promozioni (CRM Commerciale)")
    st.markdown("Archivio delle simulazioni salvate. Filtra per cliente o stato per recuperare le trattative passate.")
    
    conn = sqlite3.connect(DB_FILE)
    
    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        filtro_stato = st.selectbox("Filtra per Stato", ["Tutti", "Confermata", "Proposta"])
    with col_f2:
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT gruppo_macro FROM storico_promo ORDER BY gruppo_macro")
        gruppi_storico = ["Tutti"] + [r[0] for r in cursor.fetchall()]
        filtro_gruppo = st.selectbox("Filtra per Gruppo", gruppi_storico)
    with col_f3:
        if filtro_gruppo != "Tutti":
            cursor.execute("SELECT DISTINCT associato_insegna FROM storico_promo WHERE gruppo_macro=? ORDER BY associato_insegna", (filtro_gruppo,))
            insegne_storico = ["Tutte"] + [r[0] for r in cursor.fetchall()]
        else:
            insegne_storico = ["Tutte"]
        filtro_insegna = st.selectbox("Filtra per Insegna", insegne_storico)

    query = """
        SELECT id, data_salvataggio, stato_promo, gruppo_macro, associato_insegna, descrizione_commerciale, 
               sell_in_dal, sell_in_al, sell_out_dal, sell_out_al,
               listino_r, sconto_z, sconto_aa, min_net_net_g, net_net_am, net_net_post_promo, 
               volumi_stimati, costo_totale_extra, note 
        FROM storico_promo WHERE 1=1
    """
    params = []
    
    if filtro_stato != "Tutti":
        query += " AND stato_promo = ?"
        params.append(filtro_stato)
    if filtro_gruppo != "Tutti":
        query += " AND gruppo_macro = ?"
        params.append(filtro_gruppo)
    if filtro_insegna != "Tutte":
        query += " AND associato_insegna = ?"
        params.append(filtro_insegna)
        
    query += " ORDER BY data_salvataggio DESC"
    
    df_storico = pd.read_sql_query(query, conn, params=params)
    
    st.dataframe(
        df_storico, 
        use_container_width=True, 
        hide_index=True,
        column_config={
            "id": "ID",
            "data_salvataggio": st.column_config.DatetimeColumn("Data Salvataggio", format="DD/MM/YYYY HH:mm"),
            "stato_promo": "Stato",
            "gruppo_macro": "Gruppo",
            "associato_insegna": "Insegna",
            "descrizione_commerciale": "Prodotto",
            "sell_in_dal": st.column_config.DateColumn("Inizio Sell-In", format="DD/MM/YYYY"),
            "sell_in_al": st.column_config.DateColumn("Fine Sell-In", format="DD/MM/YYYY"),
            "sell_out_dal": st.column_config.DateColumn("Inizio Sell-Out", format="DD/MM/YYYY"),
            "sell_out_al": st.column_config.DateColumn("Fine Sell-Out", format="DD/MM/YYYY"),
            "listino_r": st.column_config.NumberColumn("Listino R", format="€ %.2f"),
            "sconto_z": st.column_config.NumberColumn("Sc. Z (%)", format="%.2f %%"),
            "sconto_aa": st.column_config.NumberColumn("Sc. AA (€)", format="€ %.2f"),
            "min_net_net_g": st.column_config.NumberColumn("Min Net Net (G)", format="€ %.3f"),
            "net_net_am": st.column_config.NumberColumn("Net Net (AM)", format="€ %.3f"),
            "net_net_post_promo": st.column_config.NumberColumn("Net Net Post-Promo", format="€ %.3f"),
            "volumi_stimati": "Volumi (Pz)",
            "costo_totale_extra": st.column_config.NumberColumn("Costo Extra", format="€ %.2f"),
            "note": "Note"
        }
    )
    
    col_export, col_delete = st.columns(2)
    
    with col_export:
        if not df_storico.empty:
            buffer_storico = io.BytesIO()
            with pd.ExcelWriter(buffer_storico, engine='openpyxl') as writer:
                df_storico.to_excel(writer, index=False, sheet_name="Storico_Promo")
                
            st.download_button(
                label="SCARICA ESTRAZIONE STORICO (Excel)",
                data=buffer_storico.getvalue(),
                file_name=f"Storico_Promozioni_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("Nessuna promozione salvata corrisponde ai filtri selezionati.")
            
    with col_delete:
        st.markdown("### Gestione Record")
        with st.expander("Elimina Promozione Salvata"):
            if not df_storico.empty:
                id_to_delete = st.selectbox("Seleziona l'ID della promozione da eliminare:", df_storico['id'].tolist())
                if st.button("ELIMINA DEFINITIVAMENTE", type="primary"):
                    try:
                        cursor.execute("DELETE FROM storico_promo WHERE id=?", (id_to_delete,))
                        conn.commit()
                        st.success(f"Promozione ID {id_to_delete} eliminata con successo.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Errore durante l'eliminazione: {e}")
            else:
                st.write("Nessun record disponibile per l'eliminazione.")
        
    conn.close()

# ==========================================
# SCHEDA 2: DATI ANAGRAFICI (PRODOTTI E LOGISTICA)
# ==========================================
elif menu == "Dati Anagrafici (Logistica)":
    st.title("Dati Anagrafici - Prodotti e Logistica")
    st.markdown("Gestione dell'anagrafica prodotti (Dati SAP e Logistici). I margini finanziari sono gestiti separatamente.")
    
    conn = sqlite3.connect(DB_FILE)
    st.subheader("Modifica Diretta Anagrafica Master")
    
    df_prodotti = pd.read_sql_query("""
        SELECT ean, codice_sap, tipo_olio, descrizione_sap, descrizione_commerciale, formato_lt, confezione,
               pezzi_cartone, cartoni_strato, strati_pallet, cartoni_pallet, conservazione_mesi, shelf_life_mesi
        FROM anagrafica_master
    """, conn)
    
    edited_prod_df = st.data_editor(
        df_prodotti, 
        num_rows="dynamic", 
        use_container_width=True,
        hide_index=True,
        key="prod_data_editor"
    )
    
    if st.button("SALVA MODIFICHE ANAGRAFICA"):
        cursor = conn.cursor()
        try:
            with conn:
                cursor.execute("DELETE FROM anagrafica_master")
                for _, r in edited_prod_df.iterrows():
                    def check_nan_float(val):
                        return float(val) if (pd.notna(val) and str(val).strip() != "") else 0.0
                    def check_nan_int(val):
                        return int(float(val)) if (pd.notna(val) and str(val).strip() != "") else 0
                    
                    cursor.execute("""
                    INSERT INTO anagrafica_master (
                        ean, codice_sap, tipo_olio, descrizione_sap, descrizione_commerciale, formato_lt, confezione,
                        pezzi_cartone, cartoni_strato, strati_pallet, cartoni_pallet, conservazione_mesi, shelf_life_mesi
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        str(r.get("ean")).strip(),
                        str(r.get("codice_sap")).strip(),
                        str(r.get("tipo_olio")).strip(),
                        str(r.get("descrizione_sap")).strip(),
                        str(r.get("descrizione_commerciale")).strip(),
                        check_nan_float(r.get("formato_lt")),
                        str(r.get("confezione")).strip(),
                        check_nan_int(r.get("pezzi_cartone")),
                        check_nan_int(r.get("cartoni_strato")),
                        check_nan_int(r.get("strati_pallet")),
                        check_nan_int(r.get("cartoni_pallet")),
                        check_nan_int(r.get("conservazione_mesi")),
                        check_nan_int(r.get("shelf_life_mesi"))
                    ))
            st.success("VERDE (APPROVATO) - L'anagrafica prodotti è stata aggiornata correttamente.")
            st.rerun()
        except Exception as e:
            st.error(f"ROSSO (BLOCCATO) - Errore durante il salvataggio dei prodotti: {e}")

    st.markdown("---")
    col_p1, col_p2 = st.columns(2)
    
    with col_p1:
        st.subheader("Esportazione Anagrafica")
        st.markdown("Scarica l'anagrafica logistica attuale per lavorarla in Excel.")
        
        buffer_prod_export = io.BytesIO()
        with pd.ExcelWriter(buffer_prod_export, engine='openpyxl') as writer:
            df_prodotti.to_excel(writer, index=False, sheet_name="Anagrafica_SAP")
            
        st.download_button(
            label="Scarica Anagrafica Prodotti (Excel)",
            data=buffer_prod_export.getvalue(),
            file_name=f"Anagrafica_Prodotti_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    with col_p2:
        st.subheader("Importazione Massiva SAP")
        st.markdown("Carica il file Excel per aggiornare l'anagrafica logistica e i guardrail finanziari.")
        uploaded_prod_file = st.file_uploader("Trascina il file Excel Anagrafica (.xlsx)", type=["xlsx"])
        
        if uploaded_prod_file is not None:
            if st.button("Conferma Scrittura Anagrafica"):
                try:
                    df_prod_import = pd.read_excel(uploaded_prod_file)
                    
                    col_map = {
                        "EAN": "ean",
                        "Codice Articolo": "codice_sap",
                        "TIPO OLIO contenuto": "tipo_olio",
                        "Descrizione articolo in SAP": "descrizione_sap",
                        "Descrizione Articolo": "descrizione_commerciale",
                        "Formato (lt)": "formato_lt",
                        "Tipologia Confezione": "confezione",
                        "Pezzi x\nCartone": "pezzi_cartone",
                        "Pezzi x Cartone": "pezzi_cartone",
                        "Cartoni x\nStrato": "cartoni_strato",
                        "Cartoni x Strato": "cartoni_strato",
                        "Strati x Pallet": "strati_pallet",
                        "Cartoni X Pallet": "cartoni_pallet",
                        "Conservazione (Mese)": "conservazione_mesi",
                        "SHELF LIFE (mesi)": "shelf_life_mesi",
                        "Margine Minimo G": "min_net_net_g",
                        "MIN_NET_NET_G": "min_net_net_g",
                        "Soglia Sicurezza G (Euro)": "min_net_net_g"
                    }
                    
                    df_prod_import = df_prod_import.rename(columns=col_map)
                    df_prod_import.columns = [str(c).lower().strip() for c in df_prod_import.columns]
                    
                    if "ean" not in df_prod_import.columns:
                        st.error("ROSSO (BLOCCATO) - Colonna 'EAN' mancante nel file Excel.")
                    else:
                        cursor = conn.cursor()
                        righe_inserite = 0
                        
                        with conn:
                            for idx, row in df_prod_import.iterrows():
                                ean_val = str(row.get("ean", "")).split('.')[0].zfill(13)
                                if not ean_val or ean_val == "0000000000000" or ean_val == "nan":
                                    continue
                                    
                                tipo_olio_raw = str(row.get("tipo_olio", "")).upper().strip()
                                if tipo_olio_raw == "EXTRA":
                                    tipo_olio_raw = "EXTRAVERGINE"

                                def get_float(col_name, default=0.0):
                                    val = row.get(col_name)
                                    if pd.isna(val) or str(val).strip() == "": return default
                                    try: return float(str(val).replace(',', '.'))
                                    except: return default
                                    
                                def get_int(col_name, default=0):
                                    val = row.get(col_name)
                                    if pd.isna(val) or str(val).strip() == "": return default
                                    try: return int(float(val))
                                    except: return default

                                min_g = row.get("min_net_net_g")
                                if pd.isna(min_g) or str(min_g).strip() == "":
                                    cursor.execute("SELECT min_net_net_g FROM guardrail_aziendali WHERE ean=?", (ean_val,))
                                    res_min = cursor.fetchone()
                                    min_g = res_min[0] if res_min else 0.0
                                else:
                                    min_g = float(str(min_g).replace(',', '.'))

                                cursor.execute("""
                                INSERT OR REPLACE INTO anagrafica_master (
                                    ean, codice_sap, tipo_olio, descrizione_sap, descrizione_commerciale, formato_lt, confezione,
                                    pezzi_cartone, cartoni_strato, strati_pallet, cartoni_pallet, conservazione_mesi, shelf_life_mesi
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """, (
                                    ean_val,
                                    str(row.get("codice_sap", "")).split('.')[0],
                                    tipo_olio_raw,
                                    str(row.get("descrizione_sap", "")),
                                    str(row.get("descrizione_commerciale", "")),
                                    get_float("formato_lt"),
                                    str(row.get("confezione", "")),
                                    get_int("pezzi_cartone"),
                                    get_int("cartoni_strato"),
                                    get_int("strati_pallet"),
                                    get_int("cartoni_pallet"),
                                    get_int("conservazione_mesi"),
                                    get_int("shelf_life_mesi")
                                ))

                                cursor.execute("""
                                INSERT OR REPLACE INTO guardrail_aziendali (ean, min_net_net_g) VALUES (?, ?)
                                """, (ean_val, min_g))
                                
                                righe_inserite += 1
                        st.success(f"VERDE (APPROVATO) - Elaborati {righe_inserite} prodotti nell'anagrafica con relativi guardrail.")
                        st.rerun()
                except Exception as e:
                    st.error(f"ROSSO (BLOCCATO) - Errore durante l'elaborazione del file: {e}")

    conn.close()

# ==========================================
# SCHEDA 3: BACK-OFFICE CONTRATTI E GUARDRAIL
# ==========================================
elif menu == "Back-Office (Contratti)":
    st.title("Back-Office - Contratti e Guardrail Finanziari")
    conn = sqlite3.connect(DB_FILE)
    
    tab_contratti, tab_guardrail = st.tabs(["Gestione Contratti GDO", "GESTIONE MINIMI NET NET"])
    
    with tab_contratti:
        st.subheader("Modifica Diretta dei Contratti in Database (A caldo)")
        
        df_database_editor = pd.read_sql_query("""
            SELECT a.id, a.gruppo_macro, a.sottogruppo, a.associato_insegna, a.livello, a.chiave_livello,
                   CASE 
                        WHEN a.livello = 'REFERENZA' THEN p.descrizione_commerciale 
                        WHEN a.livello = 'CATEGORIA' THEN 'Categoria: ' || a.chiave_livello
                        ELSE 'Contratto Quadro'
                   END as descrizione_prodotto,
                   a.listino_r,
                   a.sconto_1, a.sconto_2, a.sconto_3, a.sconto_4, a.sconto_5, a.sconto_6, a.sconto_7, a.sconto_y,
                   a.sconto_carico, a.sconto_pagamento, a.voce_contratto_1, a.voce_contratto_2, a.voce_contratto_3,
                   a.voce_contratto_4, a.voce_contratto_5
            FROM accordi_commerciali a
            LEFT JOIN anagrafica_master p ON a.chiave_livello = p.ean AND a.livello = 'REFERENZA'
        """, conn)
        
        edited_df = st.data_editor(
            df_database_editor, 
            num_rows="dynamic", 
            use_container_width=True,
            hide_index=True,
            disabled=["descrizione_prodotto"],
            key="db_data_editor"
        )
        
        if st.button("SALVA MODIFICHE CONTRATTI"):
            cursor = conn.cursor()
            try:
                with conn:
                    cursor.execute("DELETE FROM accordi_commerciali")
                    for _, r in edited_df.iterrows():
                        def check_nan(val):
                            return float(val) if (pd.notna(val) and str(val).strip() != "") else None
                        
                        cursor.execute("""
                        INSERT OR REPLACE INTO accordi_commerciali (
                            id, gruppo_macro, sottogruppo, associato_insegna, livello, chiave_livello, listino_r,
                            sconto_1, sconto_2, sconto_3, sconto_4, sconto_5, sconto_6, sconto_7, sconto_y,
                            sconto_carico, sconto_pagamento, voce_contratto_1, voce_contratto_2, voce_contratto_3,
                            voce_contratto_4, voce_contratto_5
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            check_nan(r.get("id")),
                            str(r.get("gruppo_macro")).upper().strip() if pd.notna(r.get("gruppo_macro")) else "",
                            str(r.get("sottogruppo")).upper().strip() if pd.notna(r.get("sottogruppo")) else "",
                            str(r.get("associato_insegna")).upper().strip() if pd.notna(r.get("associato_insegna")) else "",
                            str(r.get("livello")).upper().strip() if pd.notna(r.get("livello")) else "GRUPPO",
                            str(r.get("chiave_livello")).strip() if pd.notna(r.get("chiave_livello")) else "",
                            check_nan(r.get("listino_r")),
                            check_nan(r.get("sconto_1")), check_nan(r.get("sconto_2")), check_nan(r.get("sconto_3")),
                            check_nan(r.get("sconto_4")), check_nan(r.get("sconto_5")), check_nan(r.get("sconto_6")),
                            check_nan(r.get("sconto_7")), check_nan(r.get("sconto_y")), check_nan(r.get("sconto_carico")), check_nan(r.get("sconto_pagamento")),
                            check_nan(r.get("voce_contratto_1")), check_nan(r.get("voce_contratto_2")), check_nan(r.get("voce_contratto_3")),
                            check_nan(r.get("voce_contratto_4")), check_nan(r.get("voce_contratto_5"))
                        ))
                st.success("VERDE (APPROVATO) - I contratti sono stati aggiornati correttamente.")
                st.rerun()
            except Exception as e:
                st.error(f"ROSSO (BLOCCATO) - Errore durante l'elaborazione delle modifiche: {e}")

        st.markdown("---")
        col_b1, col_b2 = st.columns(2)
        
        with col_b1:
            st.subheader("Esportazione Contratti")
            
            query_accordi = """
            SELECT a.gruppo_macro as GRUPPO_MACRO, a.sottogruppo as SOTTOGRUPPO, a.associato_insegna as ASSOCIATO_INSEGNA,
                   a.livello as LIVELLO, a.chiave_livello as CHIAVE_LIVELLO,
                   CASE 
                        WHEN a.livello = 'REFERENZA' THEN p.descrizione_commerciale 
                        WHEN a.livello = 'CATEGORIA' THEN 'Accordo di Categoria: ' || a.chiave_livello
                        ELSE 'Contratto Quadro'
                   END as DESCRIZIONE_PRODOTTO,
                   a.listino_r as LISTINO_BASE_R,
                   a.sconto_1 as SCONTO_1, a.sconto_2 as SCONTO_2, a.sconto_3 as SCONTO_3, a.sconto_4 as SCONTO_4, a.sconto_5 as SCONTO_5,
                   a.sconto_6 as SCONTO_LOCAL_6, a.sconto_7 as SCONTO_LOCAL_7, a.sconto_y as SCONTO_CONTINUATIVO_Y,
                   a.sconto_carico as SCONTO_CARICO_LOGISTICA, a.sconto_pagamento as SCONTO_PAGAMENTO_AC,
                   a.voce_contratto_1 as PFA_VOCE_I, a.voce_contratto_2 as PFA_VOCE_II,
                   a.voce_contratto_3 as PFA_VOCE_III, a.voce_contratto_4 as PFA_VOCE_IV, a.voce_contratto_5 as PFA_VOCE_V
            FROM accordi_commerciali a
            LEFT JOIN anagrafica_master p ON a.chiave_livello = p.ean AND a.livello = 'REFERENZA'
            """
            df_accordi = pd.read_sql_query(query_accordi, conn)
            
            colonne_ordinate = [
                "GRUPPO_MACRO", "SOTTOGRUPPO", "ASSOCIATO_INSEGNA", "LIVELLO", "CHIAVE_LIVELLO", "DESCRIZIONE_PRODOTTO",
                "LISTINO_BASE_R", "SCONTO_1", "SCONTO_2", "SCONTO_3", "SCONTO_4", "SCONTO_5",
                "SCONTO_LOCAL_6", "SCONTO_LOCAL_7", "SCONTO_CONTINUATIVO_Y", "SCONTO_CARICO_LOGISTICA", "SCONTO_PAGAMENTO_AC",
                "PFA_VOCE_I", "PFA_VOCE_II", "PFA_VOCE_III", "PFA_VOCE_IV", "PFA_VOCE_V"
            ]
            df_accordi = df_accordi[colonne_ordinate]
            
            buffer_export = io.BytesIO()
            with pd.ExcelWriter(buffer_export, engine='openpyxl') as writer:
                df_accordi.to_excel(writer, index=False, sheet_name="Accordi_GDO")
                
            st.download_button(
                label="Scarica Template Contratti (Excel)",
                data=buffer_export.getvalue(),
                file_name=f"Backup_Contratti_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        with col_b2:
            st.subheader("Importazione Massiva Contratti")
            uploaded_file = st.file_uploader("Trascina il file Excel Contratti (.xlsx)", type=["xlsx"])
            
            if uploaded_file is not None:
                if st.button("Conferma Scrittura Contratti"):
                    try:
                        df_import = pd.read_excel(uploaded_file)
                        colonne_obbligatorie = ["GRUPPO_MACRO", "SOTTOGRUPPO", "ASSOCIATO_INSEGNA", "LIVELLO", "CHIAVE_LIVELLO"]
                        missing_cols = [c for c in colonne_obbligatorie if c not in df_import.columns]
                        
                        if missing_cols:
                            st.error(f"ROSSO (BLOCCATO) - Struttura Excel non valida. Colonne mancanti: {', '.join(missing_cols)}")
                        else:
                            cursor = conn.cursor()
                            righe_inserite = 0
                            
                            with conn:
                                for idx, row in df_import.iterrows():
                                    gruppo = str(row["GRUPPO_MACRO"]).upper().strip()
                                    sottogruppo = str(row["SOTTOGRUPPO"]).upper().strip() if (pd.notna(row.get("SOTTOGRUPPO")) and str(row.get("SOTTOGRUPPO")).strip() != "") else ""
                                    insegna = str(row["ASSOCIATO_INSEGNA"]).upper().strip() if (pd.notna(row.get("ASSOCIATO_INSEGNA")) and str(row.get("ASSOCIATO_INSEGNA")).strip() != "") else ""
                                    livello = str(row["LIVELLO"]).upper().strip()
                                    chiave_livello = str(row["CHIAVE_LIVELLO"]).strip() if pd.notna(row["CHIAVE_LIVELLO"]) else ""
                                    
                                    if livello == "REFERENZA" and chiave_livello:
                                        chiave_livello = str(chiave_livello).split('.')[0].zfill(13)

                                    cursor.execute("""
                                    INSERT OR IGNORE INTO clienti (gruppo_macro, sottogruppo, associato_insegna)
                                    VALUES (?, ?, ?)
                                    """, (gruppo, sottogruppo, insegna))

                                    def to_float_or_none(val):
                                        if pd.isna(val) or str(val).strip() == "":
                                            return None
                                        try: return float(val)
                                        except: return None

                                    cursor.execute("""
                                    INSERT OR REPLACE INTO accordi_commerciali (
                                        gruppo_macro, sottogruppo, associato_insegna, livello, chiave_livello, listino_r,
                                        sconto_1, sconto_2, sconto_3, sconto_4, sconto_5,
                                        sconto_6, sconto_7, sconto_y, sconto_carico, sconto_pagamento,
                                        voce_contratto_1, voce_contratto_2, voce_contratto_3, voce_contratto_4, voce_contratto_5
                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                    """, (
                                        gruppo, sottogruppo, insegna, livello, chiave_livello, to_float_or_none(row.get("LISTINO_BASE_R")),
                                        to_float_or_none(row.get("SCONTO_1")), to_float_or_none(row.get("SCONTO_2")), to_float_or_none(row.get("SCONTO_3")),
                                        to_float_or_none(row.get("SCONTO_4")), to_float_or_none(row.get("SCONTO_5")), to_float_or_none(row.get("SCONTO_LOCAL_6")),
                                        to_float_or_none(row.get("SCONTO_LOCAL_7")), to_float_or_none(row.get("SCONTO_CONTINUATIVO_Y")),
                                        to_float_or_none(row.get("SCONTO_CARICO_LOGISTICA")),
                                        to_float_or_none(row.get("SCONTO_PAGAMENTO_AC")), to_float_or_none(row.get("PFA_VOCE_I")),
                                        to_float_or_none(row.get("PFA_VOCE_II")), to_float_or_none(row.get("PFA_VOCE_III")),
                                        to_float_or_none(row.get("PFA_VOCE_IV")), to_float_or_none(row.get("PFA_VOCE_V"))
                                    ))
                                    righe_inserite += 1
                            st.success(f"VERDE (APPROVATO) - Elaborate {righe_inserite} regole commerciali nel Bunker.")
                            st.rerun()
                    except Exception as e:
                        st.error(f"ROSSO (BLOCCATO) - Errore durante l'importazione: {e}")

    with tab_guardrail:
        st.subheader("GESTIONE MINIMI NET NET")
        st.markdown("Questa tabella è isolata dall'anagrafica logistica. Modifica qui i limiti minimi di margine per ogni referenza.")
        
        df_guardrail = pd.read_sql_query("""
            SELECT g.ean, a.descrizione_commerciale, g.min_net_net_g
            FROM guardrail_aziendali g
            LEFT JOIN anagrafica_master a ON g.ean = a.ean
        """, conn)
        
        edited_guardrail = st.data_editor(
            df_guardrail, 
            num_rows="dynamic", 
            use_container_width=True,
            hide_index=True,
            disabled=["descrizione_commerciale"],
            column_config={
                "min_net_net_g": st.column_config.NumberColumn(
                    "Min Net Net (€)",
                    help="Soglia minima di margine in Euro",
                    format="€ %.2f",
                    step=0.01,
                )
            },
            key="guardrail_editor"
        )
        
        if st.button("SALVA MODIFICHE GUARDRAIL"):
            cursor = conn.cursor()
            try:
                with conn:
                    cursor.execute("DELETE FROM guardrail_aziendali")
                    for _, r in edited_guardrail.iterrows():
                        cursor.execute("""
                        INSERT INTO guardrail_aziendali (ean, min_net_net_g) VALUES (?, ?)
                        """, (str(r.get("ean")).strip(), float(r.get("min_net_net_g", 0.0))))
                st.success("VERDE (APPROVATO) - Guardrail aggiornati.")
                st.rerun()
            except Exception as e:
                st.error(f"ROSSO (BLOCCATO) - Errore salvataggio guardrail: {e}")

        st.markdown("---")
        
        col_g1, col_g2 = st.columns(2)
        
        with col_g1:
            st.subheader("Esportazione Guardrail")
            st.markdown("Scarica i limiti minimi attuali per lavorarli in Excel.")
            
            buffer_guardrail = io.BytesIO()
            with pd.ExcelWriter(buffer_guardrail, engine='openpyxl') as writer:
                df_guardrail.to_excel(writer, index=False, sheet_name="Guardrail_NetNet")
                
            st.download_button(
                label="Scarica Guardrail (Excel)",
                data=buffer_guardrail.getvalue(),
                file_name=f"Guardrail_Minimi_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        with col_g2:
            st.subheader("Importazione Massiva Guardrail")
            st.markdown("Carica un file Excel contenente le colonne **ean** e **min_net_net_g**.")
            uploaded_guardrail = st.file_uploader("Trascina il file Excel Guardrail (.xlsx)", type=["xlsx"], key="up_guardrail")
            
            if uploaded_guardrail is not None:
                if st.button("Conferma Scrittura Guardrail"):
                    try:
                        df_g_import = pd.read_excel(uploaded_guardrail)
                        df_g_import.columns = [str(c).lower().strip() for c in df_g_import.columns]
                        
                        if "ean" not in df_g_import.columns or "min_net_net_g" not in df_g_import.columns:
                            st.error("ROSSO (BLOCCATO) - Il file Excel deve contenere obbligatoriamente le colonne 'ean' e 'min_net_net_g'.")
                        else:
                            cursor = conn.cursor()
                            righe_inserite = 0
                            
                            with conn:
                                for idx, row in df_g_import.iterrows():
                                    ean_val = str(row.get("ean", "")).split('.')[0].zfill(13)
                                    if not ean_val or ean_val == "0000000000000" or ean_val == "nan":
                                        continue
                                        
                                    min_g = row.get("min_net_net_g", 0.0)
                                    try:
                                        min_g = float(str(min_g).replace(',', '.'))
                                    except:
                                        min_g = 0.0
                                        
                                    cursor.execute("""
                                    INSERT OR REPLACE INTO guardrail_aziendali (ean, min_net_net_g) VALUES (?, ?)
                                    """, (ean_val, min_g))
                                    righe_inserite += 1
                                    
                            st.success(f"VERDE (APPROVATO) - Aggiornati {righe_inserite} limiti minimi di margine.")
                            st.rerun()
                    except Exception as e:
                        st.error(f"ROSSO (BLOCCATO) - Errore durante l'importazione: {e}")

    st.markdown("---")
    st.markdown("<h3 style='color: #D32F2F;'>Sezione Pericolo (Danger Zone)</h3>", unsafe_allow_html=True)
    
    if PRODUCTION_MODE:
        st.info("Modalità Production: Ripristino demo disattivato.")
    else:
        st.warning("ATTENZIONE: Questa operazione ripristinerà il database allo stato iniziale.")
        pin_conferma = st.text_input("Per procedere digita la password di sicurezza 'RESET' in lettere maiuscole:")
        
        if st.button("ESEGUI HARD RESET DATABASE", disabled=(pin_conferma != "RESET")):
            try:
                seed_baseline_data(conn)
                st.success("VERDE (APPROVATO) - Database ripristinato allo stato iniziale.")
                st.rerun()
            except Exception as ex:
                st.error(f"ROSSO (BLOCCATO) - Errore durante il reset: {ex}")

    conn.close()
# ==========================================
# SCHEDA 4: REPORT SINTETICO (VERSIONE BENCHMARK COMPATTA SOTTOGRUPPI - BULLETPROOF)
# ==========================================
elif menu == "Report Sintetico":
    st.title("Report Sintetico e Analisi Contratti")
    st.markdown("---")
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    col_k1, col_k2, col_k3, col_k4 = st.columns(4)
    cursor.execute("SELECT COUNT(*) FROM accordi_commerciali")
    col_k1.metric("Totale Regole Attive", f"{cursor.fetchone()[0]}")
    
    cursor.execute("SELECT COUNT(*) FROM clienti WHERE attivo=1")
    col_k2.metric("Insegne Attive", f"{cursor.fetchone()[0]}")
    
    cursor.execute("SELECT AVG(listino_r) FROM accordi_commerciali WHERE listino_r IS NOT NULL AND listino_r > 0")
    avg_listino = cursor.fetchone()[0] or 0.0
    col_k3.metric("Listino Medio R", f"{avg_listino:.2f} €")
    
    cursor.execute("""
        SELECT AVG(voce_contratto_1 + COALESCE(voce_contratto_2,0) + COALESCE(voce_contratto_3,0) + 
                   COALESCE(voce_contratto_4,0) + COALESCE(voce_contratto_5,0)) 
        FROM accordi_commerciali
    """)
    avg_pfa = cursor.fetchone()[0] or 0.0
    col_k4.metric("PFA Medio off-invoice", f"{avg_pfa:.2f} %")
    
    st.markdown("---")
    
    contenitore_bench = st.container(border=True)
    with contenitore_bench:
        st.subheader("Benchmark Comparativo di Canale (Livello Sottogruppo)")
        st.markdown("Analisi strutturale delle asimmetrie commerciali. Gli sconti sono collassati per destinazione logica. In fase di test mettere Sagra Ex.v. CLassico lt1")
        
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            cursor.execute("SELECT DISTINCT tipo_olio FROM anagrafica_master ORDER BY tipo_olio")
            categorie_disponibili = [r[0] for r in cursor.fetchall()]
            cat_scelta = st.selectbox("1. Filtra per Categoria Merceologica", categorie_disponibili, key="bench_cat")
            
        with col_f2:
            cursor.execute("SELECT ean, descrizione_commerciale FROM anagrafica_master WHERE tipo_olio=? ORDER BY descrizione_commerciale", (cat_scelta,))
            prod_dict = {f"{p[1]} [{p[0]}]": (p[0], cat_scelta) for p in cursor.fetchall()}
            if prod_dict:
                prod_scelto_bench = st.selectbox("2. Seleziona Referenza da Analizzare", list(prod_dict.keys()), key="bench_prod")
                ean_bench, tipo_olio_bench = prod_dict[prod_scelto_bench]
            else:
                st.warning("Nessun prodotto trovato.")
                prod_scelto_bench = None

        if prod_scelto_bench:
            cursor.execute("""
                SELECT DISTINCT gruppo_macro, sottogruppo 
                FROM accordi_commerciali 
                WHERE sottogruppo != '' AND sottogruppo IS NOT NULL
                ORDER BY gruppo_macro, sottogruppo
            """)
            sottogruppi_unici = cursor.fetchall()
            benchmark_data = []
            
            for g_macro, s_gruppo in sottogruppi_unici:
                cursor.execute("""
                    SELECT associato_insegna FROM accordi_commerciali
                    WHERE gruppo_macro=? AND sottogruppo=? AND livello='REFERENZA' AND chiave_livello=? AND associato_insegna != ''
                    LIMIT 1
                """, (g_macro, s_gruppo, ean_bench))
                res_ins = cursor.fetchone()
                
                if not res_ins:
                    cursor.execute("""
                        SELECT associato_insegna FROM accordi_commerciali
                        WHERE gruppo_macro=? AND sottogruppo=? AND associato_insegna != ''
                        LIMIT 1
                    """, (g_macro, s_gruppo))
                    res_ins = cursor.fetchone()
                
                insegna_campione = res_ins[0] if res_ins else ""
                
                contratto_risolto = HierarchyResolver.resolve(conn, g_macro, s_gruppo, insegna_campione, ean_bench, tipo_olio_bench)
                
                if contratto_risolto.listino_r is not None:
                    cursor.execute("SELECT min_net_net_g FROM guardrail_aziendali WHERE ean=?", (ean_bench,))
                    res_g = cursor.fetchone()
                    soglia_g = res_g[0] if res_g else 0.0
                    
                    input_strutturale = PricingInput(
                        listino_r=contratto_risolto.listino_r,
                        sconto_1=contratto_risolto.sconto_1, sconto_2=contratto_risolto.sconto_2, sconto_3=contratto_risolto.sconto_3,
                        sconto_4=contratto_risolto.sconto_4, sconto_5=contratto_risolto.sconto_5, sconto_6=contratto_risolto.sconto_6, sconto_7=contratto_risolto.sconto_7,
                        sconto_y=contratto_risolto.sconto_y, sconto_z=Decimal("0.00"), sconto_aa=Decimal("0.00"),
                        sconto_carico=contratto_risolto.sconto_carico, sconto_pagamento=contratto_risolto.sconto_pagamento,
                        voce_i=contratto_risolto.voce_i, voce_ii=contratto_risolto.voce_ii, voce_iii=contratto_risolto.voce_iii, voce_iv=contratto_risolto.voce_iv, voce_v=contratto_risolto.voce_v,
                        min_net_net_g=Decimal(str(soglia_g))
                    )
                    calcolo_strutturale = PricingEngine.calculate(input_strutturale)
                    
                    stringa_s1_s3 = f"{float(contratto_risolto.sconto_1 or 0):.1f}% / {float(contratto_risolto.sconto_2 or 0):.1f}% / {float(contratto_risolto.sconto_3 or 0):.1f}%"
                    stringa_s4_s5 = f"{float(contratto_risolto.sconto_4 or 0):.1f}% / {float(contratto_risolto.sconto_5 or 0):.1f}%"
                    stringa_s6 = f"{float(contratto_risolto.sconto_6 or 0):.1f}%"
                    stringa_s7_y = f"S7:{float(contratto_risolto.sconto_7 or 0):.1f}% + Y:{float(contratto_risolto.sconto_y or 0):.1f}%"
                    stringa_oneri = f"Log:{float(contratto_risolto.sconto_carico or 0):.1f}% / Pag:{float(contratto_risolto.sconto_pagamento or 0):.1f}%"
                    
                    benchmark_data.append({
                        "Gruppo GDO": g_macro,
                        "Sottogruppo": s_gruppo,
                        "Origine Accordo": contratto_risolto.livello_risolto,
                        "Listino R (€)": float(contratto_risolto.listino_r),
                        "Gruppo (S1-S3)": stringa_s1_s3,
                        "Sottogruppo (S4-S5)": stringa_s4_s5,
                        "Categoria (S6)": stringa_s6,
                        "Referenza (S7+Y)": stringa_s7_y,
                        "Oneri (AB/AC)": stringa_oneri,
                        "Contratto Unificato (%)": float(calcolo_strutturale.contratto_tot_pfa),
                        "Net Net Base AM (€)": float(calcolo_strutturale.net_net_finale)
                    })
            
            if benchmark_data:
                df_out = pd.DataFrame(benchmark_data).sort_values(by="Net Net Base AM (€)")
                st.dataframe(df_out, use_container_width=True, hide_index=True)
            else:
                st.info("Nessun accordo strutturato trovato per i filtri selezionati.")

    st.markdown("---")
    st.subheader("Sintesi Dinamica e Analisi per Canale GDO")
    
    query_sintesi = """
        SELECT gruppo_macro as [Gruppo Macro],
               COUNT(*) as [Totale Righe],
               ROUND(AVG(listino_r), 2) as [Listino Medio (Euro)],
               ROUND(AVG(sconto_1), 2) as [Sconto 1 Medio (%)],
               ROUND(AVG(sconto_2), 2) as [Sconto 2 Medio (%)],
               ROUND(AVG(sconto_carico), 2) as [Oneri Logistica (%)],
               ROUND(AVG(sconto_pagamento), 2) as [Oneri Pagamento (%)],
               ROUND(AVG(voce_contratto_1 + COALESCE(voce_contratto_2,0) + COALESCE(voce_contratto_3,0) + 
                         COALESCE(voce_contratto_4,0) + COALESCE(voce_contratto_5,0)), 2) as [PFA Totale Off-Invoice (%)]
        FROM accordi_commerciali
        GROUP BY gruppo_macro
        ORDER BY [Totale Righe] DESC
    """
    df_sintesi = pd.read_sql_query(query_sintesi, conn)
    st.dataframe(df_sintesi, use_container_width=True, hide_index=True)
    
    st.markdown("---")
    st.subheader("Generatore ed Esportazione Report Consolidato di Sintesi")
    
    col_ex1, col_ex2 = st.columns(2)
    with col_ex1:
        cursor.execute("SELECT DISTINCT gruppo_macro FROM clienti WHERE attivo=1 ORDER BY gruppo_macro")
        gruppi_report = [r[0] for r in cursor.fetchall()]
        grp_rep_sel = st.selectbox("Seleziona Gruppo Macro per Esportazione", gruppi_report, key="rep_grp")
    with col_ex2:
        cursor.execute("SELECT DISTINCT associato_insegna FROM clienti WHERE gruppo_macro=? AND attivo=1 ORDER BY associato_insegna", (grp_rep_sel,))
        associati_report = [r[0] for r in cursor.fetchall()]
        ass_rep_sel = st.selectbox("Seleziona Insegna Locale per Esportazione", associati_report, key="rep_ass")
        
    if st.button("GENERA E COMPILA REPORT CONSOLIDATO EXCEL"):
        cursor.execute("SELECT sottogruppo FROM clienti WHERE gruppo_macro=? AND associato_insegna=? LIMIT 1", (grp_rep_sel, ass_rep_sel))
        res_sub = cursor.fetchone()
        sub_rep_sel = res_sub[0] if res_sub else ""
        
        cursor.execute("""
            SELECT a.ean, a.descrizione_commerciale, a.tipo_olio, COALESCE(g.min_net_net_g, 0.0), a.codice_sap, a.formato_lt, a.confezione 
            FROM anagrafica_master a
            LEFT JOIN guardrail_aziendali g ON a.ean = g.ean
        """)
        all_prods = cursor.fetchall()
        
        rows_report = []
        for p in all_prods:
            p_ean, p_desc, p_tipo, p_min_g, p_sap, p_form, p_conf = p
            resolved = HierarchyResolver.resolve(conn, grp_rep_sel, sub_rep_sel, ass_rep_sel, p_ean, p_tipo)
            
            if resolved.listino_r is not None:
                input_calc = PricingInput(
                    listino_r=resolved.listino_r,
                    sconto_1=resolved.sconto_1, sconto_2=resolved.sconto_2, sconto_3=resolved.sconto_3,
                    sconto_4=resolved.sconto_4, sconto_5=resolved.sconto_5, sconto_6=resolved.sconto_6, sconto_7=resolved.sconto_7,
                    sconto_y=resolved.sconto_y, sconto_z=Decimal("0.00"), sconto_aa=Decimal("0.00"),
                    sconto_carico=resolved.sconto_carico, sconto_pagamento=resolved.sconto_pagamento,
                    voce_i=resolved.voce_i, voce_ii=resolved.voce_ii, voce_iii=resolved.voce_iii, voce_iv=resolved.voce_iv, voce_v=resolved.voce_v,
                    min_net_net_g=Decimal(str(p_min_g))
                )
                res_calc = PricingEngine.calculate(input_calc)
                
                rows_report.append({
                    "EAN": p_ean,
                    "Codice SAP": p_sap,
                    "Descrizione Commerciale": p_desc,
                    "Formato Lt": p_form,
                    "Confezione": p_conf,
                    "Listino Base R (Euro)": float(resolved.listino_r),
                    "Sconto 1 (%)": float(resolved.sconto_1),
                    "Sconto 2 (%)": float(resolved.sconto_2),
                    "Sconto Local 6 (%)": float(resolved.sconto_6),
                    "Oneri Logistica (%)": float(resolved.sconto_carico),
                    "Oneri Pagamento (%)": float(resolved.sconto_pagamento),
                    "Prezzo Netto AF (Euro)": float(res_calc.netto_in_fattura_2),
                    "Premi Off-Invoice AL (%)": float(res_calc.contratto_tot_pfa),
                    "Prezzo Net Net AM (Euro)": float(res_calc.net_net_finale),
                    "Soglia Sicurezza G (Euro)": float(p_min_g),
                    "Delta Margine (Euro)": float(res_calc.delta_vs_min),
                    "Stato Approvazione": "VERDE" if res_calc.guardrail_ok else "ROSSO"
                })
        
        if not rows_report:
            st.warning("ATTENZIONE: Nessuna referenza in assortimento trovata per questo cliente nel database.")
        else:
            df_rep_out = pd.DataFrame(rows_report)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Consolidato_Marginalita"
            ws.views.sheetView[0].showGridLines = True
            
            font_title = Font(name="Arial", size=14, bold=True, color="FFFFFF")
            font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            font_data = Font(name="Arial", size=9)
            font_alert = Font(name="Arial", size=9, bold=True, color="9C0006")
            font_ok = Font(name="Arial", size=9, bold=True, color="006100")
            
            fill_title = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
            fill_header = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            fill_alert = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            fill_ok = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin', color='E0E0E0'),
                right=Side(style='thin', color='E0E0E0'),
                top=Side(style='thin', color='E0E0E0'),
                bottom=Side(style='thin', color='E0E0E0')
            )
            
            ws.merge_cells('A1:Q1')
            ws['A1'] = f"SALOV S.p.A. - REPORT SINTETICO CONSOLIDATO: {ass_rep_sel} ({grp_rep_sel})"
            ws['A1'].font = font_title
            ws['A1'].fill = fill_title
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 35
            
            headers = list(df_rep_out.columns)
            for col_num, h_text in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num, value=h_text)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[3].height = 25
            
            for row_num, row_data in enumerate(df_rep_out.values, 4):
                ws.row_dimensions[row_num].height = 18
                for col_num, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=val)
                    cell.font = font_data
                    cell.border = thin_border
                    
                    if col_num in [1, 2]:
                        cell.alignment = Alignment(horizontal="center")
                        cell.number_format = '@'
                    elif col_num in [4, 6, 12, 14, 15, 16]:
                        cell.alignment = Alignment(horizontal="right")
                        cell.number_format = '#,##0.000 €'
                    elif col_num in [7, 8, 9, 10, 11, 13]:
                        cell.alignment = Alignment(horizontal="right")
                        cell.number_format = '0.00" %"'
                    
                    if col_num == 17:
                        cell.alignment = Alignment(horizontal="center")
                        if val == "VERDE":
                            cell.font = font_ok
                            cell.fill = fill_ok
                        else:
                            cell.font = font_alert
                            cell.fill = fill_alert
                            
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
            
            buffer_rep = io.BytesIO()
            wb.save(buffer_rep)
            
            st.download_button(
                label=f"SCARICA REPORT EXCEL SINTESI CONTRATTO {ass_rep_sel}",
                data=buffer_rep.getvalue(),
                file_name=f"Sintesi_Contratto_{ass_rep_sel}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    conn.close()

# ==========================================
# SCHEDA 5: GUIDA OPERATIVA (VERSIONE ESTESA & AVANZATA)
# ==========================================
else:
    st.title("Manuale d'Istruzione")
    st.markdown("### Guida per la Gestione della Marginalità Salov")
    st.markdown("---")
    
    with st.expander("1. IL MOTORE DI PRICING: La Cascata Sequenziale (Esempio Numerico)", expanded=True):
        st.markdown("""
        Il simulatore non esegue mai la somma algebrica degli sconti (es. 10% + 5% non fa 15%). Il calcolo segue una **cascata geometrica sequenziale** in cui ogni sconto si applica sul risultato del passaggio precedente.
        
        Ecco un esempio reale di scomposizione per capire come si passa dal Listino al Prezzo Net Net (AM):
        
        #### Esempio Pratico: 1 Cartone di Sagra Extra Vergine Classico 1L
        *   **LISTINO BASE (R):** **10,00 €**
        *   **Sconto 1 (10,00%):** Rimane **9,00 €** *(Calcolo: 10,00 - 10%)*
        *   **Sconto 2 (5,00%):** Rimane **8,55 €** *(Calcolo: 9,00 - 5%)*
        *   **Sconto Continuativo Y (2,00%):** Rimane **8,379 €** *(Calcolo: 8,55 - 2%)*
        *   **Sconto Promozionale Z (10,00%):** Rimane **7,541 €** *(Calcolo: 8,379 - 10%)*
        *   **Sconto Taglio Prezzo Secco [AA] (0,10 €/Pz):** Rimane **7,441 €** *(Detrazione netta in Euro)*
        *   **Oneri Logistica / Pagamento (AB + AC - es. 1,5% + 1% = 2,5%):** Rimane **7,255 €** ➔ **Questo è il Netto In Fattura 2 (AF)**.
        
        #### La Fase "Off-Invoice" (Fuori Fattura)
        Sul valore di **7,255 € (AF)** si applicano i Premi Fine Anno (PFA Voci I-V) pattuiti con la Centrale. 
        Se il totale dei PFA è del **5,00%**, il sistema calcola la trattenuta finale:
        *   *Calcolo:* $7,255 \\times (1 - 0,05) = 6,892 €$
        *   **PREZZO NET NET FINALE (AM):** **6,89 €**
        
        > **Regola:** Se questo 6,89 € scende anche solo di un centesimo sotto la soglia di sicurezza **del minimo NET NET** registrata nel Back-Office per quell'EAN, l'applicazione avvisa ** BLOCCATO **.
        """)
        
    with st.expander("2. LA GERARCHIA DEI CONTRATTI: La Regola del 'Livello Superiore Comanda'", expanded=False):
        st.markdown("""
        L'applicazione utilizza un motore di risoluzione a 5 livelli. A differenza dei sistemi tradizionali, qui vige la regola del **Blocco Gerarchico (Top-Down)**: se un livello superiore definisce uno sconto, i livelli inferiori NON possono sovrascriverlo.
        
        #### I 5 Livelli (dal più forte al più debole):
        1. **GRUPPO MACRO** (es. *COOP ITALIA*) ➔ Le regole impostate qui sono blindate (Accordo Quadro). Nessun livello sottostante può modificarle.
        2. **SOTTOGRUPPO** (es. *COOP ITALIA SOTTOGRUPPO*) ➔ Può aggiungere sconti solo se il Gruppo Macro ha lasciato la cella vuota. Non può essere sovrascritto dai livelli inferiori.
        3. **CATEGORIA** (es. *EXTRAVERGINE*) ➔ Comanda su Insegna e Referenza, ma subisce le regole di Gruppo e Sottogruppo.
        4. **ASSOCIATO / INSEGNA** (es. *ALLEANZA 3.0*) ➔ Regole locali. Non possono sovrascrivere la Categoria o i Gruppi.
        5. **REFERENZA (EAN)** ➔ Il livello più basso. Definisce il Listino Base (R) e sconti specifici solo se nessun livello superiore li ha già bloccati.
        
        #### Come gestire i campi in Tabella (Casi Reali):
        
        *   **Caso A: Il Blocco della Centrale (Nessuna Sovrascrittura)**
            Se il Gruppo COOP ITALIA fissa lo Sconto 1 al **10%**, anche se per sbaglio inserisci 15% sulla singola Referenza, il sistema ignorerà il 15% e manterrà il 10%. Il livello superiore vince sempre.
            
        *   **Caso B: L'Ereditarietà (La Cella Vuota)**
            Se il Gruppo lascia lo Sconto 6 vuoto (NULL), il Sottogruppo o l'Insegna sono liberi di valorizzarlo. Il primo livello (partendo dall'alto) che inserisce un valore, lo blocca per tutti i livelli sottostanti.
            
        *   **Caso C: Il Fuori Assortimento**
            Il Listino Base (R) si inserisce quasi sempre al livello 5 (Referenza). Se manca, il prodotto risulta "Fuori Assortimento" e non può essere simulato.
        """)

    with st.expander("3. LE DUE MODALITÀ DI LAVORO: Target vs Manuale Spot", expanded=False):
        st.markdown("""
        Nella scheda principale puoi scegliere due modi diversi di attaccare il pricing a seconda di cosa stai discutendo con il buyer della GDO:
        
        ####  Modalità A: Partenza da Prezzo Target (Consigliata)
        La usi quando il buyer ti dice: *"Voglio vendere la bottiglia a scaffale a questo prezzo, quindi a te la pago esattamente X"*.
        1. Seleziona la modalità **A**.
        2. Inserisci nel campo il prezzo richiesto dal cliente.
        3. Il motore calcola istantaneamente al millesimo lo **Sconto Promozionale Z (%)** necessario per arrivare a quel prezzo.
        4. Se il target inserito fa scendere la marginalità sotto la soglia di sicurezza, il sistema calcolerà comunque lo sconto ma ti avviserà del blocco.
        
        ####  Modalità B: Tentativi Spot Manuali (Uso Libero)
        La usi per fare simulazioni classiche o per testare scenari "Cosa succede se...".
        1. Seleziona la modalità **B**.
        2. Muovi manualmente lo Sconto Promozionale Z o lo Sconto AA.
        3. Tieni d'occhio i campi **Sconto Promo MAX Consentito [Z]** e **Sconto Unitario MAX Consentito [AA]**: ti indicano esattamente fino a dove puoi spingerti con la percentuale o con l'Euro secco prima che il semaforo passi da Verde a Rosso.
        """)

    with st.expander("4. BACK-OFFICE ED EXCEL: Come Aggiornare i Dati in Sicurezza", expanded=False):
        st.markdown("""
        L'applicazione si alimenta con i dati reali delle anagrafiche e dei contratti. Puoi fare manutenzione in due modi:
        
        ####  Variante 1: Modifiche rapide "a caldo" direttamente a schermo
        1. Vai su **Back-Office (Contratti)** o **Dati Anagrafici**.
        2. Fai doppio clic sulla cella che vuoi modificare all'interno della griglia dati.
        3. Digita il nuovo valore (es. cambia un listino o modifica un PFA).
        4. Clicca sul pulsante **SALVA MODIFICHE** per rendere la modifica operativa immediatamente su tutto il simulatore.
        
        ####  Variante 2: Caricamento Massivo in Excel (Operazioni Pesanti)
        Se devi aggiornare l'intero piano contrattuale annuale:
        1. Clicca su **Scarica Template Contratti (Excel)** per avere il backup completo del database attuale.
        2. Lavora i dati comodamente sul tuo Excel aziendale.
        3. ** ATTENZIONE AI CODICI EAN ** Excel tende a trasformare i codici a 13 cifre in numeri scientifici (es. `800221E+12`). Prima di salvare, assicurati che la colonna **EAN** e **CHIAVE_LIVELLO** siano formattate esplicitamente come **TESTO**, altrimenti l'importazione corromperà l'anagrafica impedendo al simulatore di riconoscere i prodotti.
        4. Trascina il file salvato nel box di importazione e clicca su **Conferma Scrittura**.
        """)
